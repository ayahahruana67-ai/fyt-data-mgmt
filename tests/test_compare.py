# -*- coding: utf-8 -*-
"""表格比对引擎回归测试(合成簿,可移植)。

钉死:按关键列配对(乱序仍对上)、值归一化('10'==10)、
只在A/只在B、重复键报告、表头自动识别、导出报告落盘。
"""
import os
import tempfile
import unittest
import warnings

import openpyxl

from core import compare_core as C

warnings.filterwarnings("ignore", message="Workbook contains no default style")


class _Tmp(unittest.TestCase):
    def setUp(self):
        self._tmp = tempfile.mkdtemp(prefix="fyt_cmp_")

    def tearDown(self):
        import shutil
        shutil.rmtree(self._tmp, ignore_errors=True)

    def mk(self, name, rows, pre=0):
        """rows[0] 为表头;pre 为表头前插入的空/杂行数(测表头识别)。"""
        p = os.path.join(self._tmp, name)
        wb = openpyxl.Workbook(); ws = wb.active
        for _ in range(pre):
            ws.append([])
        for row in rows:
            ws.append(row)
        wb.save(p)
        return p


HDR = ["物料编码", "名称", "数量"]


class TestCompare(_Tmp):
    def _cmp(self, rows_a, rows_b, key="物料编码", columns=None):
        ha, ra = C.read_table(self.mk("a.xlsx", [HDR] + rows_a))
        hb, rb = C.read_table(self.mk("b.xlsx", [HDR] + rows_b))
        return C.compare(ha, ra, hb, rb, key, columns=columns)

    def test_identical_no_diff(self):
        rows = [["M01", "纸箱", 10], ["M02", "螺丝", 5]]
        r = self._cmp(rows, [list(x) for x in rows])
        self.assertEqual(r["counts"]["diffs"], 0)
        self.assertEqual(r["counts"]["only_a"], 0)
        self.assertEqual(r["counts"]["only_b"], 0)

    def test_single_cell_change(self):
        r = self._cmp([["M01", "纸箱", 10]], [["M01", "纸箱", 12]])
        self.assertEqual(r["counts"]["diffs"], 1)
        self.assertEqual(r["diffs"][0]["column"], "数量")
        self.assertEqual(r["diffs"][0]["a"], 10)
        self.assertEqual(r["diffs"][0]["b"], 12)

    def test_reordered_still_matches(self):
        # 行顺序不同,按关键列仍应全部对上、0 差异
        a = [["M01", "纸箱", 10], ["M02", "螺丝", 5]]
        b = [["M02", "螺丝", 5], ["M01", "纸箱", 10]]
        r = self._cmp(a, b)
        self.assertEqual(r["counts"]["diffs"], 0)
        self.assertEqual(r["counts"]["matched"], 2)

    def test_numeric_text_equivalence(self):
        # A 存数字 10,B 存文本 "10" / " 10 " -> 不算差异
        r = self._cmp([["M01", "纸箱", 10]], [["M01", "纸箱", " 10 "]])
        self.assertEqual(r["counts"]["diffs"], 0)

    def test_numeric_key_matches_across_types(self):
        # 关键列一边是整数 10、另一边是公式算出的 10.0 -> 应对上,不误判"只在单边"
        r = self._cmp([[10, "纸箱", 1]], [[10.0, "纸箱", 1]])
        self.assertEqual(r["counts"]["only_a"], 0)
        self.assertEqual(r["counts"]["only_b"], 0)
        self.assertEqual(r["counts"]["matched"], 1)

    def test_text_code_key_not_collapsed(self):
        # 文本编码键:"001" 与 "1" 是不同键,前导零有意义,不得折叠成同键
        r = self._cmp([["001", "甲", 1]], [["1", "乙", 1]])
        self.assertEqual(sorted(o["key"] for o in r["only_a"]), ["001"])
        self.assertEqual(sorted(o["key"] for o in r["only_b"]), ["1"])

    def test_only_in_each_side(self):
        r = self._cmp([["M01", "x", 1], ["M02", "y", 2]],
                      [["M02", "y", 2], ["M03", "z", 3]])
        self.assertEqual([o["key"] for o in r["only_a"]], ["M01"])
        self.assertEqual([o["key"] for o in r["only_b"]], ["M03"])

    def test_duplicate_key_reported(self):
        r = self._cmp([["M01", "x", 1], ["M01", "x", 9]], [["M01", "x", 1]])
        self.assertIn("M01", r["dup_a"])

    def test_header_autodetect_with_preamble(self):
        # 表头前有空行,仍应识别到真正表头
        p = self.mk("pre.xlsx", [HDR, ["M01", "纸箱", 10]], pre=3)
        headers, rows = C.read_table(p)
        self.assertEqual(headers, HDR)
        self.assertEqual(len(rows), 1)

    def test_common_columns(self):
        self.assertEqual(
            C.common_columns(["a", "b", "c"], ["b", "c", "d"]), ["b", "c"])

    def test_key_must_exist(self):
        with self.assertRaises(ValueError):
            self._cmp([["M01", "x", 1]], [["M01", "x", 1]], key="不存在列")

    def test_export_report_creates_file(self):
        r = self._cmp([["M01", "x", 1], ["M02", "y", 2]],
                      [["M01", "x", 9], ["M03", "z", 3]])
        path = C.export_report(r, out_dir=self._tmp)
        self.assertTrue(os.path.exists(path))
        wb = openpyxl.load_workbook(path)
        self.assertEqual(set(wb.sheetnames),
                         {"概要", "差异明细", "只在A", "只在B"})
        wb.close()

    def test_run_end_to_end(self):
        a = self.mk("out_a.xlsx", [HDR, ["M01", "纸箱", 10], ["M02", "螺丝", 5]])
        b = self.mk("out_b.xlsx", [HDR, ["M01", "纸箱", 10], ["M02", "螺丝", 6]])
        logs = []
        r = C.run(a, b, key="物料编码", out_dir=self._tmp, log=logs.append)
        self.assertEqual(r["counts"]["diffs"], 1)
        self.assertTrue(os.path.exists(r["report_path"]))
        self.assertTrue(any("比对完成" in l for l in logs))


if __name__ == "__main__":
    unittest.main()


def _qt_available():
    try:
        os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
        import PySide2  # noqa
        return True
    except Exception:
        return False


@unittest.skipUnless(_qt_available(), "PySide2 不可用,跳过 UI 冒烟测试")
class TestCompareUI(_Tmp):
    """比对页 UI 冒烟:页面能建、选文件后关键列下拉自动填、结果弹窗能建。
    不进事件循环、不跑子线程,只验证控件装配与数据流。"""

    @classmethod
    def setUpClass(cls):
        from PySide2.QtWidgets import QApplication
        cls.app = QApplication.instance() or QApplication([])

    def test_page_and_dialog_build(self):
        from ui.pages.compare_page import ComparePage
        from ui.dialogs.compare_review import CompareResultPanel

        class _FakeMain:                     # 页面只用到 self.main,给个占位
            pass
        pg = ComparePage(_FakeMain())
        self.assertFalse(pg.panel.run_btn.isEnabled())   # 无文件时禁用

        a = self.mk("a.xlsx", [HDR, ["M01", "x", 1], ["M02", "y", 2]])
        b = self.mk("b.xlsx", [HDR, ["M02", "y", 9], ["M03", "z", 3]])
        pg.z_a.set_paths([a]); pg.z_b.set_paths([b])
        pg._refresh()
        keys = [pg.cmb_key.itemText(i) for i in range(pg.cmb_key.count())]
        self.assertEqual(keys, HDR)                       # 公共列填入下拉
        self.assertTrue(pg.panel.run_btn.isEnabled())     # 齐全后启用

        res = C.run(a, b, key="物料编码", out_dir=self._tmp)
        dlg = CompareResultPanel(res)                     # 面板能从结果构建
        from PySide2.QtWidgets import QTabWidget
        tabs = dlg.findChild(QTabWidget)
        self.assertEqual(tabs.count(), 4)


class TestCompareOutputPath(_Tmp):
    """输出路径纳入全程序统一约定:feature 'compare' 有中文目录名。"""

    def test_feature_dir_registered(self):
        from core import paths
        self.assertIn("compare", paths.FEATURE_DIRS)
        out = paths.resolve_output_dir("compare", mode="custom",
                                       custom_root=self._tmp)
        self.assertIn("表格比对", out)          # 未注册会退化成英文 'compare'

    def test_run_default_dir_is_unified(self):
        # out_dir=None 时应走统一根(不再落 cwd);切到 custom 根验证不污染文档
        from core import settings as sm
        st = sm.get_settings()
        saved = dict(st._data)
        try:
            st._data["output_mode"] = "custom"
            st._data["custom_output_root"] = self._tmp
            a = self.mk("a.xlsx", [HDR, ["M01", "x", 1]])
            b = self.mk("b.xlsx", [HDR, ["M01", "x", 2]])
            res = C.run(a, b, key="物料编码")
            self.assertIn(self._tmp, res["out_dir"])
            self.assertIn("表格比对", res["out_dir"])
        finally:
            st._data.clear(); st._data.update(saved)
