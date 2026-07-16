# -*- coding: utf-8 -*-
"""
考勤填表核心逻辑（不依赖GUI，便于测试与复用）。
- 从“每日统计表”读取：姓名、日期、上班1打卡时间、下班1打卡时间（支持 .xlsx/.xlsm/.xls）
- 按 姓名 + 日期 匹配，填入考勤表的：上班时间（系统）、下班时间（系统）
- 自动算“实际上/下班时间”：把系统打卡时间按半小时进位(上班)/退位(下班)。
  例：上班 7:56、7:31 → 8:00；下班 8:13、8:24 → 8:00。（opts.auto_actual，默认开）
- 实际工作时间 = 下班时间（实际） - 上班时间（实际） - 休息时间
- 加班 = 实际工时 - 标准工时（可在高级选项调整；不足记0）
- 异常数据(下班早于上班、或扣休息后为负)不再按跨夜 +24 处理：该行标黄、
  工时写真实(负)值，并汇总到输出表新增子表『异常核对报告』，供人工核对。
- 未匹配到打卡数据的行(系统数据查无此人此日)也记入异常报告并标黄；
  行内已标“假/休/调休”等非工时词、或已手填实际时间算出工时的除外。
公共解析/常量/选项/输出路径统一来自 common_core。保留目标文件原有格式。
"""
import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

from . import common_core as cc
from .common_core import Options, norm_name, norm_date, parse_time, to_hours, parse_rest

# 异常行标黄填充
_HL_FILL = PatternFill(fill_type="solid", fgColor="FFFF00")

# 向后兼容：仍暴露该常量名
STANDARD_WORKDAY_HOURS = cc.STANDARD_WORKDAY_HOURS


# ---------- 读取源表（每日统计表） ----------
def _detect_source_header(rows, opts, path=""):
    """在一个 sheet 里定位 姓名/日期/上班1打卡/下班1打卡 表头。
    返回 (hdr_idx0, cols) 或 None。
    手动列映射(opts.resolve_roles)优先；其次按 header_row 限定行；否则自动识别。"""
    roles = opts.resolve_roles(path)
    header = opts.resolve_header(path)
    if roles and all(k in roles for k in ("name", "date", "on", "off")):
        hdr0 = (header - 1) if header else 0
        return hdr0, {k: roles[k] for k in ("name", "date", "on", "off")}
    cand = [header - 1] if header else range(min(6, len(rows)))
    for i in cand:
        if i < 0 or i >= len(rows):
            continue
        row = rows[i]
        joined = [norm_name(x) for x in row]
        if "姓名" not in joined or not any("上班1打卡" in x for x in joined):
            continue
        cols = {}
        for c, val in enumerate(joined):
            if val == "姓名": cols["name"] = c
            elif val == "日期": cols["date"] = c
            elif "上班1打卡" in val: cols["on"] = c
            elif "下班1打卡" in val: cols["off"] = c
        if all(k in cols for k in ("name", "date", "on", "off")):
            return i, cols
    return None


def load_source(path, opts=None):
    """读取单个每日统计表 -> {(姓名,(y,m,d)):(上班打卡,下班打卡)}。支持多子表，取首个命中表头的子表。"""
    opts = opts or cc.DEFAULTS
    sheets = cc.read_sheets(path)
    want_sheet = opts.resolve_sheet(path)
    if want_sheet:
        sheets = [(n, r) for (n, r) in sheets if n == want_sheet]
        if not sheets:
            raise ValueError("文件 %s 中找不到工作表 '%s'" % (os.path.basename(path), want_sheet))
    ds_override = opts.resolve_data_start(path)
    for sname, rows in sheets:
        det = _detect_source_header(rows, opts, path)
        if det is None:
            continue
        hdr, cols = det
        data = {}
        start = (ds_override - 1) if ds_override else (hdr + 1)
        for r in range(start, len(rows)):
            row = rows[r]
            def g(c):
                return row[c] if c < len(row) else None
            name = norm_name(g(cols["name"]))
            d = norm_date(g(cols["date"]))
            if not name or d is None:
                continue
            on = g(cols["on"]); off = g(cols["off"])
            on_s = "" if on is None else str(on).strip()
            off_s = "" if off is None else str(off).strip()
            data[(name, d)] = (on_s, off_s)
        return data
    raise ValueError("源表 %s 未找到表头（需含 '姓名' 和 '上班1打卡时间'）。" % os.path.basename(path))


def load_source_multi(paths, opts=None, log=None):
    """读取并合并多个每日统计表。重复(姓名+日期)按 opts.conflict 处理。
    返回 (data, stat)：stat={"files","records","conflicts"}"""
    opts = opts or cc.DEFAULTS
    log = log or (lambda *a, **k: None)
    if isinstance(paths, str):
        paths = [paths]
    merged = {}; conflicts = 0
    for p in paths:
        try:
            one = load_source(p, opts)
        except Exception as e:
            log("  · [跳过] %s（读取失败：%s）" % (os.path.basename(p), e)); continue
        log("  · [读取] %s：%d 条打卡记录" % (os.path.basename(p), len(one)))
        for key, (new_on, new_off) in one.items():
            if key not in merged:
                merged[key] = (new_on, new_off); continue
            conflicts += 1
            old_on, old_off = merged[key]
            if opts.conflict == "first":
                pass                                  # 保留先者
            elif opts.conflict == "warn":
                log("    ! 重复且不覆盖：%s %s" % (key[0], "-".join(map(str, key[1]))))
            else:                                     # last：非空才覆盖
                use_on = new_on if (new_on and new_on not in ("-", "—")) else old_on
                use_off = new_off if (new_off and new_off not in ("-", "—")) else old_off
                merged[key] = (use_on, use_off)
    if conflicts:
        cn = {"last": "后者覆盖", "first": "先者优先", "warn": "不覆盖仅提示"}
        log("  注意：%d 条(姓名+日期)重复，按【%s】处理。" % (conflicts, cn.get(opts.conflict)))
    return merged, {"files": len(paths), "records": len(merged), "conflicts": conflicts}


# ---------- 异常行标黄 + 报告子表 ----------
def _highlight_row(ws, r, cols):
    """把该行涉及的数据列(从最小列到最大列)整段标黄，便于人工定位。"""
    used = [c for c in cols.values() if c]
    if not used:
        return
    for c in range(min(used), max(used) + 1):
        ws.cell(r, c).fill = _HL_FILL


def _write_report_sheet(wb, anomalies):
    """在工作簿追加“异常核对报告”子表，逐条列出异常。anomalies 为空则不建表。"""
    if not anomalies:
        return
    title = "异常核对报告"
    if title in wb.sheetnames:                 # 避免重名冲突
        del wb[title]
    ws = wb.create_sheet(title)
    head = ["工作表", "行号", "姓名", "日期", "班次", "实际上班", "实际下班",
            "休息时间", "算出工时", "异常原因"]
    ws.append(head)
    for c in range(1, len(head) + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True)
        cell.fill = _HL_FILL
        cell.alignment = Alignment(horizontal="center")
    for a in anomalies:
        ws.append([a["sheet"], a["row"], a["name"], a["date"], a.get("shift", ""),
                   a["act_on"], a["act_off"], a["rest"], a["hours"], a["reason"]])
    widths = [14, 6, 10, 12, 8, 10, 10, 10, 10, 34]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"


def _row_rest_mark(ws, r, cols, skip):
    """行内若已标注“假/休/调休”等非工时词，返回该词；否则返回 ""。
    检查上/下班(系统/实际)四个时间格——休息日通常在这里写“休”“假”。"""
    for k in ("sys_on", "act_on", "sys_off", "act_off"):
        c = cols.get(k)
        if not c:
            continue
        txt = norm_name(ws.cell(r, c).value)
        if txt in skip and txt != "":
            return txt
    return ""


def compute_shift(a_on, a_off, rest, opts):
    """按两班制算工时并判异常。
    a_on/a_off : datetime.time（实际上/下班，已进退位）；rest：休息小时。
    返回 (hours, reason, shift, base)：
      hours——实际工作时间(小时,四舍五入2位,异常时为真实值可负)；
      reason——异常原因(空串=正常)；shift——"白班"/"夜班"；base——该班标准工时(加班基准)。
    规则：上班钟点 ≥ opts.night_start_hour 判夜班；夜班下班<上班时 +24 修正。
    异常：白班下班早于上班；任意班扣休息后为负；夜班时长超上限(疑漏打卡)。
    """
    on_h = to_hours(a_on)
    off_h = to_hours(a_off)
    is_night = opts.night_shift and on_h >= opts.night_start_hour
    shift = "夜班" if is_night else "白班"
    base = opts.night_workday_hours if is_night else opts.workday_hours
    raw = off_h - on_h
    reason = ""
    if is_night:
        if raw < 0:                     # 跨零点：+24 修正为真实时长
            raw += 24
        hours = round(raw - rest, 2)
        if raw > opts.night_max_hours:  # 时长过长，疑似漏打卡
            reason = "夜班时长 %.2fh 超上限 %.2fh（疑漏打卡，请核对）" % (raw, opts.night_max_hours)
        elif hours < 0:
            reason = "夜班工时不足扣休息（时长 %.2fh 少于休息 %.2fh）" % (raw, rest)
    else:
        hours = round(raw - rest, 2)
        if raw < 0:
            reason = "下班早于上班（实际下班 %s 早于实际上班 %s）" % (
                cc.fmt_time(a_off), cc.fmt_time(a_on))
        elif hours < 0:
            reason = "工时不足扣休息（时长 %.2fh 少于休息 %.2fh）" % (raw, rest)
    return hours, reason, shift, base


# ---------- 填写目标表（保留原格式，用 openpyxl 写回） ----------
def find_target_columns(ws, opts=None, path=""):
    """在目标表定位所需列，返回 (header_row, cols)。
    手动列映射(opts.resolve_roles)优先；其次 header_row 指定行；否则第1行自动识别。
    cols 内的列号统一为 1-based（openpyxl）；手动映射存的是 0-based，取用时 +1。"""
    opts = opts or cc.DEFAULTS
    hr = opts.resolve_header(path) or 1
    header = {norm_name(ws.cell(hr, c).value).replace("\n", ""): c
              for c in range(1, ws.max_column + 1)}
    def col(*keys):
        for k in keys:
            if k in header:
                return header[k]
        return None
    cols = {
        "name": col("姓名"), "date": col("日期"),
        "sys_on": col("上班时间（系统）"), "act_on": col("上班时间（实际）"),
        "sys_off": col("下班时间（系统）"), "act_off": col("下班时间（实际）"),
        "rest": col("休息时间"), "work": col("实际工作时间"), "ot": col("加班"),
    }
    roles = opts.resolve_roles(path)     # 手动映射覆盖对应角色（0-based -> 1-based）
    for k, c0 in roles.items():
        if k in cols:
            cols[k] = c0 + 1
    return hr, cols


def fill_workbook(target_path, source_data, out_path, opts=None, log=None):
    """把 source_data 填入目标表所有工作表（或指定表），算工时/加班，另存 out_path。返回统计 dict。"""
    opts = opts or cc.DEFAULTS
    log = log or (lambda *a, **k: None)
    skip = opts.skip_set()                     # “假/休/调休”等非工时标记，不当未匹配异常
    wb = openpyxl.load_workbook(target_path)   # 不用 data_only，保留格式
    stats = {"sheets": [], "matched": 0, "filled_time": 0, "computed_work": 0,
             "unmatched": 0, "filled_actual": 0, "anomalies": 0}
    anomalies = []                     # 跨所有子表汇总,最后写“异常核对报告”
    sheets = wb.worksheets
    want_sheet = opts.resolve_sheet(target_path)
    if want_sheet:
        sheets = [w for w in sheets if w.title == want_sheet]
        if not sheets:
            wb.close()
            raise ValueError("目标表中找不到工作表 '%s'" % want_sheet)
    ds_override = opts.resolve_data_start(target_path)
    for ws in sheets:
        hr, cols = find_target_columns(ws, opts, target_path)
        if not cols["name"] or not cols["date"]:
            log("跳过工作表 '%s'（未找到姓名/日期列）" % ws.title); continue
        s_matched = s_filled = s_work = s_unmatched = s_actual = s_anomaly = 0
        start = ds_override if ds_override else (hr + 1)
        for r in range(start, ws.max_row + 1):
            name = norm_name(ws.cell(r, cols["name"]).value)
            d = norm_date(ws.cell(r, cols["date"]).value)
            if not name or d is None:
                continue
            key = (name, d)
            matched = key in source_data
            if matched:
                on_s, off_s = source_data[key]; s_matched += 1
                if cols["sys_on"] and on_s and on_s not in ("-", "—"):
                    ws.cell(r, cols["sys_on"]).value = on_s; s_filled += 1
                if cols["sys_off"] and off_s and off_s not in ("-", "—"):
                    ws.cell(r, cols["sys_off"]).value = off_s
                # 自动算“实际上/下班时间”：上班进位、下班退位到最近半小时，填入实际列
                if opts.auto_actual:
                    if cols["act_on"]:
                        t = cc.round_half_hour(parse_time(on_s), "up")
                        if t is not None:
                            ws.cell(r, cols["act_on"]).value = cc.fmt_time(t); s_actual += 1
                    if cols["act_off"]:
                        t = cc.round_half_hour(parse_time(off_s), "down")
                        if t is not None:
                            ws.cell(r, cols["act_off"]).value = cc.fmt_time(t)
            else:
                s_unmatched += 1
            # 实际工作时间 = 下班(实际) - 上班(实际) - 休息
            computed = False
            if cols["work"] and cols["act_on"] and cols["act_off"]:
                a_on = parse_time(ws.cell(r, cols["act_on"]).value)
                a_off = parse_time(ws.cell(r, cols["act_off"]).value)
                if a_on is not None and a_off is not None:
                    computed = True
                    rest = parse_rest(ws.cell(r, cols["rest"]).value) if cols["rest"] else 0.0
                    hours, reason, shift, base = compute_shift(a_on, a_off, rest, opts)
                    ws.cell(r, cols["work"]).value = hours; s_work += 1
                    if reason:
                        _highlight_row(ws, r, cols)
                        s_anomaly += 1
                        anomalies.append({
                            "sheet": ws.title, "row": r, "name": name,
                            "date": "%04d-%02d-%02d" % d, "shift": shift,
                            "act_on": cc.fmt_time(a_on), "act_off": cc.fmt_time(a_off),
                            "rest": rest, "hours": hours, "reason": reason})
                        log("    ! 异常：%s 第%d行 %s(%s) —— %s"
                            % (ws.title, r, name, shift, reason))
                    elif opts.overtime and cols["ot"]:        # 正常行才算加班(按各班基准)
                        ot = round(hours - base, 2)
                        ws.cell(r, cols["ot"]).value = ot if ot > 0 else 0
            # 未匹配到打卡、又没算出工时：记为异常（休息日“假/休”等除外）
            if not matched and not computed:
                mark = _row_rest_mark(ws, r, cols, skip)
                if not mark:
                    _highlight_row(ws, r, cols)
                    s_anomaly += 1
                    anomalies.append({
                        "sheet": ws.title, "row": r, "name": name,
                        "date": "%04d-%02d-%02d" % d, "shift": "-",
                        "act_on": "", "act_off": "", "rest": "", "hours": "",
                        "reason": "未匹配到打卡数据（系统数据中查无此人此日，或姓名/日期不一致、缺卡）"})
                    log("    ! 异常：%s 第%d行 %s —— 未匹配到打卡数据" % (ws.title, r, name))
        stats["sheets"].append((ws.title, s_matched, s_filled, s_work, s_unmatched))
        for k, v in (("matched", s_matched), ("filled_time", s_filled),
                     ("computed_work", s_work), ("unmatched", s_unmatched),
                     ("filled_actual", s_actual), ("anomalies", s_anomaly)):
            stats[k] += v
        log("工作表 '%s'：匹配 %d 行，填打卡 %d 处，算实际时间 %d 处，算工时 %d 行%s" %
            (ws.title, s_matched, s_filled, s_actual, s_work,
             ("，异常 %d 行(已标黄)" % s_anomaly) if s_anomaly else ""))
    _write_report_sheet(wb, anomalies)         # 有异常才追加报告子表
    if anomalies:
        log("⚠ 共 %d 行异常数据需人工核对，已标黄并汇总到子表『异常核对报告』。" % len(anomalies))
    wb.save(out_path)
    return stats


# ---------- 统一入口：与对账功能同构 ----------
def run(targets, sources, opts=None, log=None, out_dir=None):
    """考勤填报统一入口（输出方式与工时对账一致）。
    targets : 待填考勤表路径列表（或单个）
    sources : 系统数据表路径列表（打卡来源）
    out_dir : 输出目录；不传则用统一 paths 系统（文档下统一文件夹）。
    输出：out_dir 下生成 名字_已填写_时间戳.xlsx（同批次共用一个时间戳）
    返回 {"out_files":[...], "out_dir":..., "source_stat":..., "results":[(target,out,stats)]}
    """
    opts = opts or cc.DEFAULTS
    log = log or (lambda *a, **k: None)
    if isinstance(targets, str):
        targets = [targets]
    if isinstance(sources, str):
        sources = [sources]
    ts = cc.timestamp()          # 同一批次共用一个时间戳
    if out_dir is None:
        out_dir = _unified_out_dir("attendance", ts, src=targets[0]) or cc.make_out_dir(targets[0])
    log("采用选项：" + opts.summary())
    log("① 读取并合并系统数据（%d 个文件）..." % len(sources))
    data, sstat = load_source_multi(sources, opts=opts, log=log)
    log("   合并后共 %d 条打卡记录。" % sstat["records"])

    out_files, results = [], []
    for i, tgt in enumerate(targets, 1):
        log("\n② 填写第 %d/%d 个待填表：%s" % (i, len(targets), os.path.basename(tgt)))
        base = os.path.splitext(os.path.basename(tgt))[0]
        op = cc.out_path(out_dir, base, "_已填写", ".xlsx", ts=ts)
        stats = fill_workbook(tgt, data, op, opts=opts, log=log)
        log("   匹配 %d 行、填打卡 %d 处、算工时 %d 行、未匹配 %d 行、异常 %d 行"
            % (stats["matched"], stats["filled_time"], stats["computed_work"],
               stats["unmatched"], stats["anomalies"]))
        log("   已保存：%s" % op)
        out_files.append(op); results.append((tgt, op, stats))
    return {"out_files": out_files, "out_dir": out_dir,
            "source_stat": sstat, "results": results}


def _unified_out_dir(feature, ts=None, src=None):
    """通过统一 paths/settings 解析输出目录；导入失败则回退到原逻辑。
    src: beside 模式下用于定位源文件目录。"""
    try:
        from . import paths as _paths
        from . import settings as _settings
        st = _settings.get_settings()
        kw = st.output_kwargs()
        if src and not kw.get("src_path"):
            kw["src_path"] = src
        return _paths.resolve_output_dir(feature, ts=ts, **kw)
    except Exception:
        return None
