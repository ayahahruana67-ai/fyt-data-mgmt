# -*- coding: utf-8 -*-
"""文件预览读取 —— 只取前若干行,喂给侧栏预览网格。

供常驻侧栏的"文件预览"分区用:选中/点击一个文件即在右侧看到它的前 N 行,
无需打开 Excel。刻意轻量:.xlsx/.xlsm 用 read_only 流式只读前 N 行即停,
不载入整簿;.xls 走 xlrd;.csv 探测编码/分隔符读前 N 行。绝不改动源文件。

返回结构统一为 PreviewData,UI 只管渲染,不关心来源格式。
兼容 Windows 7 + Python 3.8 + PySide2。
"""
import os
import csv
import datetime

DEFAULT_ROWS = 30          # 预览行数(含表头行)
DEFAULT_COLS = 40          # 预览列数上限,避免超宽表把网格撑爆


class PreviewData(object):
    """一张表的预览快照。rows 为二维列表(已文本化),sheets 为同簿其余表名。"""
    def __init__(self, path, sheet, rows, sheets=None, truncated=False, error=""):
        self.path = path
        self.sheet = sheet
        self.rows = rows                 # [[cell_text, ...], ...]
        self.sheets = sheets or []       # 该工作簿所有子表名(csv 为空)
        self.truncated = truncated       # 是否因行数上限被截断
        self.error = error               # 非空表示读取失败,rows 为空

    @property
    def ncols(self):
        return max((len(r) for r in self.rows), default=0)

    @property
    def nrows(self):
        return len(self.rows)


def _fmt(v):
    """单元格文本化:None->''、日期->ISO、浮点整数去 .0、其余 str。"""
    if v is None:
        return ""
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, (datetime.datetime, datetime.date)):
        try:
            return v.strftime("%Y-%m-%d")
        except Exception:
            return str(v)
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def list_sheets(path):
    """列出工作簿子表名(供预览面板切换)。csv/.xls 读失败返回 []。"""
    ext = os.path.splitext(path)[1].lower()
    try:
        if ext in (".xlsx", ".xlsm"):
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True)
            try:
                return list(wb.sheetnames)
            finally:
                wb.close()
        if ext == ".xls":
            import xlrd
            return list(xlrd.open_workbook(path, on_demand=True).sheet_names())
    except Exception:
        pass
    return []


def read_preview(path, sheet=None, max_rows=DEFAULT_ROWS, max_cols=DEFAULT_COLS):
    """读取 path(可指定 sheet)的前 max_rows 行 × max_cols 列,返回 PreviewData。

    出错不抛异常,错误信息写进 PreviewData.error,让 UI 就地显示而非崩溃。"""
    ext = os.path.splitext(path)[1].lower()
    try:
        if not os.path.isfile(path):
            return PreviewData(path, sheet, [], error="文件不存在")
        if ext in (".xlsx", ".xlsm"):
            return _read_xlsx(path, sheet, max_rows, max_cols)
        if ext == ".xls":
            return _read_xls(path, sheet, max_rows, max_cols)
        if ext in (".csv", ".txt"):
            return _read_csv(path, max_rows, max_cols)
        return PreviewData(path, sheet, [], error="不支持预览的类型:%s" % ext)
    except Exception as e:
        return PreviewData(path, sheet, [], error=str(e))


def _read_xlsx(path, sheet, max_rows, max_cols):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        names = list(wb.sheetnames)
        ws = wb[sheet] if (sheet and sheet in names) else wb[names[0]]
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows:
                break
            rows.append([_fmt(c) for c in row[:max_cols]])
        # read_only 下无法廉价知道总行数,以"取满即可能被截断"近似
        truncated = len(rows) >= max_rows
        return PreviewData(path, ws.title, rows, sheets=names, truncated=truncated)
    finally:
        wb.close()


def _read_xls(path, sheet, max_rows, max_cols):
    import xlrd
    book = xlrd.open_workbook(path)
    names = book.sheet_names()
    sh = book.sheet_by_name(sheet) if (sheet and sheet in names) else book.sheet_by_index(0)
    rows = []
    for r in range(min(sh.nrows, max_rows)):
        rows.append([_fmt(sh.cell(r, c).value) for c in range(min(sh.ncols, max_cols))])
    return PreviewData(path, sh.name, rows, sheets=list(names),
                       truncated=sh.nrows > max_rows)


def _read_csv(path, max_rows, max_cols):
    # 探测编码:优先 utf-8-sig(带 BOM 的国产导出常见),回退 gbk
    with open(path, "rb") as fb:
        raw = fb.read(4096)
    enc = "utf-8-sig"
    try:
        raw.decode("utf-8-sig")
    except Exception:
        enc = "gbk"
    rows = []
    with open(path, "r", encoding=enc, errors="replace", newline="") as f:
        sample = f.read(2048)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
        except Exception:
            dialect = csv.excel
        reader = csv.reader(f, dialect)
        for i, row in enumerate(reader):
            if i >= max_rows:
                break
            rows.append([_fmt(c) for c in row[:max_cols]])
    return PreviewData(path, "", rows, truncated=len(rows) >= max_rows)
