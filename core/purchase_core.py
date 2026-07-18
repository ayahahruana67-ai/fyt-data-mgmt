# -*- coding: utf-8 -*-
"""
采购数对账 —— 核心逻辑
======================
把"我方对账单"与"供应商对单明细"逐行匹配，生成三样产物：
  · 两张原表副本，数量列上色（绿=对上，黄=未对上）；
  · 对账汇报单：已对上并排 + 双方各自未对上 + 未对上原因，供人工一眼核对。

匹配规则（经真实数据验证，与人工上色结果 168/168 完全一致）：
  · 必需：材料名称 + 规格 + 数量 三者一致（规格忽略大小写/空格/*×；数量按数值比）；
  · 材料编号：两边都有则必须相等，且批次不得矛盾（容忍"漏打/多打一位"笔误）；
  · 批次号：仅当某侧缺编号时作兜底判据，并作优先级评分让"编号+批次都吻合"先锁定；
  · 一对一贪心配对：高分优先，每行只用一次。

表头行与列位置自动识别，不写死列号。兼容 Windows 7 + Python 3.8 + PySide2。
"""
import os
import re

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import paths as _paths
from . import settings as settings_mod
from .common_core import warn_if_uncached   # 公式未刷新检测(读关键表前告警)
from . import header_detect
from . import shape_detect

# 表头关键词 -> 角色。识别时先精确匹配，再包含匹配；每列只归一个角色。
HEADER_KEYS = {
    "no":    ["材料编号", "物料编号", "存货编码", "产品编号", "编号", "料号"],
    "name":  ["材料名称", "物料名称", "存货名称", "产品名称", "名称", "品名"],
    "spec":  ["规格型号", "规格", "型号"],
    "unit":  ["计量单位", "单位"],
    "qty":   ["采购数量", "对账数量", "结算数量", "数量合计", "数量"],
    "batch": ["批次号", "生产批次", "批次", "批号"],
    "note":  ["备注", "说明"],
}
GREEN = PatternFill("solid", fgColor="FF92D050")    # 对上
YELLOW = PatternFill("solid", fgColor="FFFFFF00")   # 未对上


def norm_name(v):
    return "" if v is None else str(v).strip().replace(" ", "")


def norm_spec(v):
    if v is None:
        return ""
    return str(v).strip().upper().replace(" ", "").replace("*", "X").replace("×", "X")


def norm_no(v):
    """材料编号归一(参考 delivery_core.norm_code): 让 "123.0"/"00123"/"123" 归为同值。
    否则 float 编码 123.0->"123.0"、文本 "00123" 原样, 会导致同编号匹配不上误报"对方无此编号"。"""
    if v is None:
        return ""
    # float 整数取 int(去掉尾随 .0)
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    s = str(v).strip().upper()
    # 纯数字串去前导零(00123->123); 保留全零为 "0", 非纯数字(含字母/横杠)原样
    if s.isdigit():
        s = s.lstrip("0") or "0"
    return s


def batch_core(v):
    """批次核心：去掉开头连续英文字母前缀(如 GKMYR/GMVNR)，便于跨供应商比较。"""
    if v is None:
        return ""
    s = str(v).strip().upper()
    s2 = re.sub(r"^[A-Z]+", "", s)
    return s2 or s


def batch_compat(a, b):
    """批次是否兼容：核心相等或互为子串（应对 26021 与 26021-02 这类粒度差异）。"""
    ca, cb = batch_core(a), batch_core(b)
    return bool(ca and cb and (ca == cb or ca in cb or cb in ca))


def _one_indel(a, b):
    """是否仅差一次"插入/删除"(长度差恰为1，短串=长串删一个字符)。

    只认漏打/多打一位这类笔误(26004↔2604)，
    不认等长替换——一位数字被改写(21004↔21014、26004↔26014)几乎必是不同批次。
    """
    la, lb = len(a), len(b)
    if abs(la - lb) != 1:
        return False
    if la > lb:                        # 让 a 为较短者
        a, b = b, a
    for k in range(len(b)):            # 短串 = 长串删一个字符？
        if a == b[:k] + b[k + 1:]:
            return True
    return False


def _split_batch(v):
    """把批次核心拆成 (前缀, 子批次尾号)，如 26004-01 -> ('26004','01')。"""
    core = batch_core(v)
    if "-" in core:
        p, s = core.rsplit("-", 1)
        return p, s
    return core, ""


def batch_consistent(a, b):
    """编号已相等时，批次是否可信为"同一行"。

    只容忍供应商"漏打/多打一位"的笔误(26004↔2604)，拒绝真正不同的批次：
      · 子批次尾号两侧都有且不同 → 不同批，拒绝(26004-01 vs 26004-02)；
      · 前缀被改写一位(等长替换)→ 不同批，拒绝(21004-01 vs 21014-01)；
      · 前缀差异超过一次增删 → 不同批，拒绝(26004-01 vs 26010-01)。
    """
    if batch_compat(a, b):
        return True
    pa, sa = _split_batch(a)
    pb, sb = _split_batch(b)
    if sa and sb and sa != sb:
        return False
    if not pa or not pb:
        return True
    return _one_indel(pa, pb)


def qty_eq(a, b):
    """数量相等：优先数值比较，非数值退化为去空字符串比较。"""
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) < 1e-9
    except (TypeError, ValueError):
        return str(a).strip() == str(b).strip()


# ---------------------------------------------------------------------------
# 表头/列自动识别
# ---------------------------------------------------------------------------
def detect_layout(ws, scan_rows=12, log=None):
    """在前若干行里找表头行并映射列。返回 (header_row, {角色:列号}) 或 (None, {})。

    选"能命中最多角色"的行为表头行；要求至少含 name 与 qty 两列，否则视为未识别。
    薄封装:算法在 header_detect.detect_layout,此处只传本功能的常量。
    采购表无供应商角色,故无需 exclude_contains;若日后加供应商列,守卫随引擎自动可用。
    """
    return header_detect.detect_layout(
        ws, HEADER_KEYS, require=("name", "qty"), scan_rows=scan_rows, log=log)


# 数据形态画像(表头识别失败时兜底):角色顺序≈典型列序,末位 bool 为必需。
_SHAPE_PROFILE = [
    ("no", shape_detect.CODE, False),
    ("name", shape_detect.TEXT, True),
    ("spec", shape_detect.TEXT, False),
    ("unit", shape_detect.TEXT, False),
    ("qty", shape_detect.NUMBER, True),
]


def detect_layout_or_shape(ws, scan_rows=12, log=None):
    """先按表头文字识别;失败再按数据形态兜底。返回 (header_row, col_map, source)。

    source: "header" / "shape"(需用户确认) / None。兜底映射不应静默落盘。"""
    hr, col = detect_layout(ws, scan_rows=scan_rows, log=log)
    if hr:
        return hr, col, "header"
    hr2, col2, _conf = shape_detect.detect_by_shape(
        ws, _SHAPE_PROFILE, scan_rows=scan_rows, log=log)
    if hr2:
        return hr2, col2, "shape"
    return None, {}, None


def load_rows(path, sheet=None, log=None):
    """读取一张表的有效数据行。返回 (rows, layout)。
    rows: [{r, no, name, spec, unit, qty, batch}]，已过滤空行/合计行。
    """
    # 数量列常是公式; 未刷新时 data_only 读出 None -> qty_eq 恒 False, 该行永远标黄。
    # 读表前先醒目告警(命中列含"数量"即提示)。
    if log:
        warn_if_uncached(path, log, sheet=sheet, what="数量")
    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
        header_row, col = detect_layout(ws, log=log)
        if not header_row:
            raise ValueError("未能在 %s / %s 中识别表头（需含“名称”“数量”列）"
                             % (os.path.basename(path), ws.title))
        rows = []
        none_qty = 0                                # 数量读出 None 的行数(疑似公式未刷新)
        for r in range(header_row + 1, ws.max_row + 1):
            nm = ws.cell(r, col["name"]).value
            if nm is None:
                continue
            if isinstance(nm, str) and ("合计" in nm or "小计" in nm or "总计" in nm):
                continue
            qv = ws.cell(r, col["qty"]).value
            if qv is None:
                none_qty += 1
            rows.append({
                "r": r,
                "no": ws.cell(r, col["no"]).value if "no" in col else None,
                "name": nm,
                "spec": ws.cell(r, col["spec"]).value if "spec" in col else None,
                "unit": ws.cell(r, col["unit"]).value if "unit" in col else None,
                "qty": qv,
                "batch": ws.cell(r, col["batch"]).value if "batch" in col else None,
            })
        layout = {"sheet": ws.title, "header_row": header_row, "col": col}
    finally:
        wb.close()                                  # 显式释放, 避免文件句柄泄漏
    # 数量取不到值的行单独提示(而非直接判不等标黄)
    if none_qty and log:
        log("⚠ 《%s》有 %d 行的数量未取到值(可能公式未刷新), 这些行将无法正确对账, 请核对。"
            % (os.path.basename(path), none_qty))
    return rows, layout


# ---------------------------------------------------------------------------
# 匹配引擎
# ---------------------------------------------------------------------------
def _can_match(a, b):
    """两行是否可配对（不含一对一约束）。"""
    if norm_name(a["name"]) != norm_name(b["name"]):
        return False
    if norm_spec(a["spec"]) != norm_spec(b["spec"]):
        return False
    if not qty_eq(a["qty"], b["qty"]):
        return False
    na, nb = norm_no(a["no"]), norm_no(b["no"])
    if na and nb:
        # 编号双方都有：必须相等，且批次不得矛盾（容忍单字符笔误，拒绝不同批次）
        return na == nb and batch_consistent(a["batch"], b["batch"])
    # 编号缺一侧或双缺: 有批次时按批次兜底
    if batch_compat(a["batch"], b["batch"]):
        return True
    # 双方均无编号且批次都空: 名称+规格+数量已全同, 退化为按此三项配对(见 pair_note 会提示)
    if not na and not nb and not batch_core(a["batch"]) and not batch_core(b["batch"]):
        return True
    return False


def _pair_score(a, b):
    """配对优先级：编号相等(+2) 与 批次兼容(+1) 叠加，越高越先锁定。"""
    na, nb = norm_no(a["no"]), norm_no(b["no"])
    s = 2 if (na and nb and na == nb) else 0
    if batch_compat(a["batch"], b["batch"]):
        s += 1
    return s


def match_rows(rows1, rows2):
    """一对一贪心配对。返回 (matched1, matched2, pairs)。
    matched1/2: 布尔列表；pairs: [(i, j, score)]。
    """
    cand = []
    for i, a in enumerate(rows1):
        for j, b in enumerate(rows2):
            if _can_match(a, b):
                cand.append((_pair_score(a, b), i, j))
    cand.sort(key=lambda x: -x[0])
    m1 = [False] * len(rows1)
    m2 = [False] * len(rows2)
    pairs = []
    for s, i, j in cand:
        if m1[i] or m2[j]:
            continue
        m1[i] = m2[j] = True
        pairs.append((i, j, s))
    return m1, m2, pairs


# ---------------------------------------------------------------------------
# 未对上原因诊断
# ---------------------------------------------------------------------------
def diagnose(row, others_all, others_avail):
    """给一条未对上的行，逐级定位到"对方对应那一条"并说明原因。

    定位顺序：编号 → 名称+规格 → 批次。关键：只在【仍未配对】的对方行里找对应，
    绝不引用已被别的行配走的数量/批次（否则会报出对方"未对上清单"里根本查不到的数字）。
    只有真正找到对应批次的未配对行时，才报"数量不符(对方数量:X)"，且仅列该批次的数量。
    """
    def same_ns(pool):
        return [b for b in pool
                if norm_name(row["name"]) == norm_name(b["name"])
                and norm_spec(row["spec"]) == norm_spec(b["spec"])]

    na = norm_no(row["no"])
    if na:
        if not [b for b in others_all if norm_no(b["no"]) == na]:
            return "对方无此编号(%s)" % row["no"]
        avno = [b for b in others_avail if norm_no(b["no"]) == na]
        if not avno:
            return "对方此编号明细已全部配对，本行多出"
        ns = same_ns(avno)
        if not ns:
            b = avno[0]
            return "同编号但名称/规格不同(对方:%s %s)" % (b["name"], b["spec"])
        bc = [b for b in ns if batch_consistent(row["batch"], b["batch"])]
        if not bc:
            return "同编号同规格但对方无此批次(%s)" % (row["batch"] or "空")
        qs = "/".join(sorted({str(b["qty"]) for b in bc}))
        return "同编号同批次但数量不符(对方数量:%s)" % qs

    ns = same_ns(others_avail)
    if not ns:
        return "对方(未对上项中)无此料/规格"
    bc = [b for b in ns if batch_compat(row["batch"], b["batch"])]
    if not bc:
        return "同名同规格但对方无此批次(%s)" % (row["batch"] or "空")
    qs = "/".join(sorted({str(b["qty"]) for b in bc}))
    return "同名同规格同批次但数量不符(对方数量:%s)" % qs


def quantity_conflicts(rows1, rows2, matched1, matched2):
    """数量不符的"近似同一行"疑点：对账最该人工看的部分。

    只在【未对上】的行里找：名称+规格一致、编号兼容(都有则相等)、批次兼容，
    仅数量不同——即"本该是同一行、只差数量"的真实差异。已配对的行不再报。
    每条只取批次最接近的一个对家，避免同料多批次刷屏。
    """
    out = []
    for i, a in enumerate(rows1):
        if matched1[i]:
            continue
        best = None
        for j, b in enumerate(rows2):
            if matched2[j]:
                continue
            if norm_name(a["name"]) != norm_name(b["name"]):
                continue
            if norm_spec(a["spec"]) != norm_spec(b["spec"]):
                continue
            na, nb = norm_no(a["no"]), norm_no(b["no"])
            if na and nb and na != nb:
                continue
            if not batch_compat(a["batch"], b["batch"]):
                continue
            if qty_eq(a["qty"], b["qty"]):
                continue
            # 批次核心完全相等的优先作为对家
            exact = batch_core(a["batch"]) == batch_core(b["batch"])
            if best is None or (exact and not best[1]):
                best = (b, exact)
        if best is not None:
            out.append((a, best[0]))
    return out


# ---------------------------------------------------------------------------
# 对账汇报单（并排）
# ---------------------------------------------------------------------------
def pair_note(a, b):
    """给一对已匹配行标注需人工留意之处；把握较大的干净匹配返回空串。

    只对"匹配依据偏弱"的行提示，避免刷屏：
      · 编号相等但批次不兼容 → 靠"漏打一位"笔误规则救回来的，务必核对；
      · 两侧都无编号、且批次核心不完全相等（仅子串近似）→ 依据较弱，建议核对。
    编号相等且批次兼容、或批次核心完全一致的，视为可靠，不加备注。
    """
    na, nb = norm_no(a["no"]), norm_no(b["no"])
    if na and nb:
        if not batch_compat(a["batch"], b["batch"]):
            return "批次疑似笔误(%s↔%s)，请核对" % (a["batch"], b["batch"])
        return ""
    # 两侧编号缺失：仅当批次核心并非完全相等时提示
    if batch_core(a["batch"]) != batch_core(b["batch"]):
        return "无编号且批次仅近似(%s↔%s)，请核对" % (a["batch"], b["batch"])
    # 双方均无编号且批次都空: 仅凭名称+规格+数量配对, 依据最弱, 务必核对
    if not batch_core(a["batch"]):
        return "无编号且无批次，仅按名称+规格+数量匹配，请核对"
    return ""


_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEAD_FILL = PatternFill("solid", fgColor="FF4472C4")
_HEAD_FONT = Font(bold=True, color="FFFFFFFF")
_SEC_FILL = PatternFill("solid", fgColor="FFD9E1F2")
_MATCH_FILL = PatternFill("solid", fgColor="FFE2EFDA")   # 淡绿：对上
_UNMATCH_FILL = PatternFill("solid", fgColor="FFFFF2CC")  # 淡黄：未对上
_FLAG_FILL = PatternFill("solid", fgColor="FFFCE4D6")     # 淡橙：需核对
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _put(ws, r, c, val, fill=None, font=None, align=None, border=True):
    cell = ws.cell(r, c, val)
    if fill:
        cell.fill = fill
    cell.font = font or Font()
    cell.alignment = align or _CENTER
    if border:
        cell.border = _BORDER
    return cell


def apply_colors(path, sheet, matched, rows, qty_col, out_path):
    """在原文件副本上给数量列上色（保留原格式/公式），另存为 out_path。"""
    wb = openpyxl.load_workbook(path)
    try:
        ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
        for flag, info in zip(matched, rows):
            ws.cell(info["r"], qty_col).fill = GREEN if flag else YELLOW
        try:
            wb.save(out_path)
            return out_path
        except PermissionError:
            raise PermissionError("无法保存 %s —— 请先在 Excel 里关闭该文件后重试" % out_path)
    finally:
        wb.close()                                  # 显式释放, 避免文件句柄泄漏


# 汇报单列布局：我方 5 列 + 供方 5 列 + 备注；共 12 列
_RH = ["序号",
       "编号", "名称", "规格", "批次", "数量",
       "编号", "名称", "规格", "批次", "数量",
       "备注"]


def _write_row(ws, r, seq, a, b, fill, note):
    """写一行并排数据；a/b 任一为 None 时该侧留空。"""
    _put(ws, r, 1, seq, fill)
    cols_a = ("no", "name", "spec", "batch", "qty")
    for k, key in enumerate(cols_a):
        _put(ws, r, 2 + k, a[key] if a else "", fill,
             align=_LEFT if key in ("name", "spec") else _CENTER)
    for k, key in enumerate(cols_a):
        _put(ws, r, 7 + k, b[key] if b else "", fill,
             align=_LEFT if key in ("name", "spec") else _CENTER)
    _put(ws, r, 12, note, _FLAG_FILL if note else fill, align=_LEFT)


def build_report(rows1, rows2, m1, m2, pairs, out_path, name1, name2):
    """生成对账汇报单：对上并排 + 各自未对上 + 原因，单张表一眼核对。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "对账汇报单"
    ncol = len(_RH)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    _put(ws, 1, 1, "采购数对账汇报单  （%s ↔ %s）" % (name1, name2),
         _HEAD_FILL, Font(bold=True, size=14, color="FFFFFFFF"))
    ws.row_dimensions[1].height = 26
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)
    _put(ws, 2, 1,
         "%s %d 行 / 对上 %d / 未对上 %d      %s %d 行 / 对上 %d / 未对上 %d      配对 %d 对"
         % (name1, len(rows1), sum(m1), len(m1) - sum(m1),
            name2, len(rows2), sum(m2), len(m2) - sum(m2), len(pairs)),
         _SEC_FILL, Font(bold=True), align=_LEFT)
    return _finish_report(ws, wb, rows1, rows2, m1, m2, pairs,
                          out_path, name1, name2, ncol)


def _section(ws, r, text, ncol):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
    _put(ws, r, 1, text, _SEC_FILL, Font(bold=True), align=_LEFT)
    return r + 1


def _header_band(ws, r):
    """两级表头：第一级分组(我方/供方/备注)，第二级字段。"""
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 1, end_column=1)
    _put(ws, r, 1, "序号", _HEAD_FILL, _HEAD_FONT)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    _put(ws, r, 2, "我方", _HEAD_FILL, _HEAD_FONT)
    ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=11)
    _put(ws, r, 7, "供方", _HEAD_FILL, _HEAD_FONT)
    ws.merge_cells(start_row=r, start_column=12, end_row=r + 1, end_column=12)
    _put(ws, r, 12, "备注", _HEAD_FILL, _HEAD_FONT)
    for c in range(2, 12):
        _put(ws, r + 1, c, _RH[c - 1], _HEAD_FILL, _HEAD_FONT)
    return r + 2


def _finish_report(ws, wb, rows1, rows2, m1, m2, pairs,
                   out_path, name1, name2, ncol):
    r = 4
    r = _section(ws, r, "一、已对上（%d 对）" % len(pairs), ncol)
    hb = _header_band(ws, r)
    ws.cell(r, 2).value = name1        # 用实际名称覆盖"我方/供方"分组名
    ws.cell(r, 7).value = name2
    r = hb
    seq = 1
    for i, j, _s in sorted(pairs, key=lambda p: rows1[p[0]]["r"]):
        a, b = rows1[i], rows2[j]
        note = pair_note(a, b)
        _write_row(ws, r, seq, a, b, _FLAG_FILL if note else _MATCH_FILL, note)
        r += 1
        seq += 1

    # 未对上：我方（诊断只在对方"仍未配对"的行里找对应）
    um1 = [row for row, ok in zip(rows1, m1) if not ok]
    um2 = [row for row, ok in zip(rows2, m2) if not ok]
    r = _section(ws, r + 1, "二、%s 未对上（%d 条）" % (name1, len(um1)), ncol)
    r = _header_band(ws, r)
    ws.cell(r - 2, 2).value = name1
    for row in um1:
        _write_row(ws, r, row["r"], row, None, _UNMATCH_FILL,
                   diagnose(row, rows2, um2))
        r += 1

    # 未对上：供方
    r = _section(ws, r + 1, "三、%s 未对上（%d 条）" % (name2, len(um2)), ncol)
    r = _header_band(ws, r)
    ws.cell(r - 2, 7).value = name2
    for row in um2:
        _write_row(ws, r, row["r"], None, row, _UNMATCH_FILL,
                   diagnose(row, rows1, um1))
        r += 1

    widths = [6, 13, 14, 16, 14, 7, 13, 14, 16, 14, 7, 30]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A4"
    try:
        wb.save(out_path)
        return out_path
    except PermissionError:
        raise PermissionError("无法保存 %s —— 请先在 Excel 里关闭该文件后重试" % out_path)


def _out_name(path):
    base, ext = os.path.splitext(os.path.basename(path))
    return "%s_对账结果%s" % (base, ext if ext.lower() == ".xlsx" else ".xlsx")


# ---------------------------------------------------------------------------
# 统一入口（与其余功能一致：run(..., out_dir=None, log=None) -> dict）
# ---------------------------------------------------------------------------
def run(file1, file2, sheet1=None, sheet2=None, name1="我方", name2="供方",
        out_dir=None, log=None):
    """采购数对账主流程。生成两张上色表 + 一张对账汇报单，返回结果 dict。

    file1/file2 : 我方对账单 / 供应商对单明细（xlsx 路径）
    name1/name2 : 报表中双方显示名称（不填用"我方/供方"，绝不写死具体供应商）
    """
    def _lg(msg):
        if log:
            log(msg)

    rows1, lay1 = load_rows(file1, sheet1, log=_lg)
    rows2, lay2 = load_rows(file2, sheet2, log=_lg)
    _lg("%s 有效行: %d（表头第%d行，工作表 %s）"
        % (name1, len(rows1), lay1["header_row"], lay1["sheet"]))
    _lg("%s 有效行: %d（表头第%d行，工作表 %s）"
        % (name2, len(rows2), lay2["header_row"], lay2["sheet"]))

    m1, m2, pairs = match_rows(rows1, rows2)
    _lg("配对成功 %d 对" % len(pairs))
    _lg("%s 对上 %d / 未对上 %d" % (name1, sum(m1), len(m1) - sum(m1)))
    _lg("%s 对上 %d / 未对上 %d" % (name2, sum(m2), len(m2) - sum(m2)))

    qc = quantity_conflicts(rows1, rows2, m1, m2)
    if qc:
        _lg("\n【数量不一致疑点】%d 处：" % len(qc))
        for a, b in qc:
            _lg("  %s行%s(量%s) ↔ %s行%s(量%s) | %s %s %s"
                % (name1, a["r"], a["qty"], name2, b["r"], b["qty"],
                   a["no"], a["name"], a["spec"]))

    if out_dir is None:
        st = settings_mod.get_settings()
        out_dir = _paths.resolve_output_dir("purchase", **st.output_kwargs())
    else:
        os.makedirs(out_dir, exist_ok=True)
    _lg("输出文件夹：%s" % out_dir)

    o1 = apply_colors(file1, lay1["sheet"], m1, rows1, lay1["col"]["qty"],
                      os.path.join(out_dir, _out_name(file1)))
    o2 = apply_colors(file2, lay2["sheet"], m2, rows2, lay2["col"]["qty"],
                      os.path.join(out_dir, _out_name(file2)))
    _lg("已生成上色表：%s" % o1)
    _lg("已生成上色表：%s" % o2)

    report = os.path.join(out_dir, "采购数对账汇报单.xlsx")
    rp = build_report(rows1, rows2, m1, m2, pairs, report, name1, name2)
    _lg("已生成汇报单：%s" % rp)
    return {"rows1": rows1, "rows2": rows2, "matched1": m1, "matched2": m2,
            "pairs": pairs, "qty_conflicts": qc, "out1": o1, "out2": o2,
            "report": rp, "out_dir": out_dir, "layout1": lay1, "layout2": lay2}
