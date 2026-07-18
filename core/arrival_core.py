# -*- coding: utf-8 -*-
"""
到料明细表核心逻辑（从 arrival_table_app.pyw 抽出，与 GUI 解耦）
================================================================
做什么：扫描"送货计划表"，识别批次号、统计未收料物料，生成
"YYYYMMDD每日主料到料明细.xlsx"（微软雅黑、蓝色标签、细边框，样式与示例一致）。

改动点（相对原程序）：
· 输出目录不再写死"程序目录/output"，改由统一 paths 系统解析（见 run）；
· 批次记忆(top_label/total/remark)改存全局 settings，不再各自散落；
· 纯逻辑，可单独测试与复用。

兼容 Windows 7 + Python 3.8。
"""
import os
import re
import glob
import datetime

import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter

from . import paths as _paths
from . import settings as _settings
from . import shape_detect
from .common_core import warn_if_uncached   # 公式未刷新检测(读关键表前告警)

# 兜底固定列(表头识别失败时回退): 编码2/名称3/供应商5/需求7/剩余未收12
COL_CODE, COL_NAME, COL_SUPPLIER, COL_DEMAND, COL_REMAIN = 2, 3, 5, 7, 12
FIRST_COL, BATCH_STRIDE, DATA_START_ROW = 3, 9, 8
DEFAULT_TOTAL = 566
DEFAULT_TOP_LABEL = "截止16点的数据"
# 输出列宽: 序号/编码/名称/供应商/需求数/剩余未收数/备注
COL_WIDTHS = [7.9, 17.6, 29.6, 34.4, 9.0, 10.5, 11.9]
HEADER_SCAN_ROWS = 8               # 表头最多出现在前几行内

# 表头别名(去空格后子串匹配)。供应商要"信息"不要"代码/编号"。
ALIAS_CODE   = ("物料编码", "物料编号", "材料编码", "材料编号", "物料号", "料号")
ALIAS_NAME   = ("物料名称", "材料名称", "品名")
ALIAS_DEMAND = ("需求数", "需求数量", "需求量", "计划需求")
ALIAS_REMAIN = ("剩余未收数", "剩余未收", "未收数", "未收", "未到货", "未到")
CONFIG_PATH = os.path.join(os.path.expanduser("~"), ".arrival_table_config.json")

# 批次号识别: 优先“订单XXX批次”, 其次文件名中 字母+数字(可带-数字段) 组合
RE_TITLE = re.compile(r'订单\s*([A-Za-z0-9]+(?:-[A-Za-z0-9]+)*)\s*批次')
RE_NAME  = re.compile(r'([A-Za-z]{2,}\d{2,}(?:-[A-Za-z0-9]+)*)')

FONT_NAME = "微软雅黑"
thin = Side(style='thin')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
FONT = Font(name=FONT_NAME, size=11)
FONT_B = Font(name=FONT_NAME, size=11, bold=True)
# 表头/标签蓝色填充: 主题色4(蓝) tint≈0.4, 与示例一致
BLUE_FILL = PatternFill(patternType='solid', fgColor=Color(theme=4, tint=0.3999755851924192))
def beijing_date():
    return (datetime.datetime.utcnow() + datetime.timedelta(hours=8)).strftime("%Y%m%d")
def _first_ws(wb):
    """返回要使用的工作表: 优先活动表, 否则第一张. 不再依赖表名 Sheet1。"""
    try:
        ws = wb.active
        if ws is not None:
            return ws
    except Exception:
        pass
    return wb[wb.sheetnames[0]]


def _pick_data_ws(wb, log=None):
    """挑出真正含到货计划数据的子表, 避免静默选错(如空的 'Sheet2')。

    规则: 优先活动表; 若活动表 locate_columns 失败(无有效表头), 再扫其余子表,
    取第一个能定位成功的; 都失败则回退到活动/第一张(交由固定列兜底)。
    单文件时行为与 _first_ws 完全一致。多子表时把选中的子表名 log 出来,不再静默。
    """
    active = _first_ws(wb)
    if len(wb.sheetnames) <= 1:
        return active
    # 活动表本身有效 -> 直接用它(尊重用户在 Excel 里停留的表)
    if locate_columns(active) is not None:
        if log:
            log("· 多子表: 读取活动表《%s》" % active.title)
        return active
    # 活动表无效: 找第一个能定位到表头的子表
    for name in wb.sheetnames:
        ws = wb[name]
        if ws is active:
            continue
        if locate_columns(ws) is not None:
            if log:
                log("· 多子表: 活动表《%s》无有效表头, 改读《%s》"
                    % (active.title, name))
            return ws
    if log:
        log("⚠ 多子表: 各子表均未识别到有效表头, 按活动表《%s》读取(可能选错文件)"
            % active.title)
    return active

def _norm(v):
    """去空白/换行后的字符串, 便于表头匹配。"""
    if v is None:
        return ""
    s = str(v)
    for ch in ("\r\n", "\n", "\r", "\t", " ", "　"):
        s = s.replace(ch, "")
    return s.strip()

def _match(text, aliases):
    return any(a in text for a in aliases)

def locate_columns(ws):
    """按表头文字灵活定位各列, 不再写死列号。

    在前几行里找"物料编码"所在行当表头行, 然后在该行按别名匹配各字段列:
      供应商——优先含'供应商信息', 否则退而取含'供应商'但**不含**'代码/编号'的列;
      剩余未收数——按 ALIAS_REMAIN 匹配。
    任一关键列(编码/需求/剩余未收)缺失, 返回 None -> 调用方回退到固定列。
    返回 dict(code,name,supplier,demand,remain,header_row)。
    """
    scan = min(HEADER_SCAN_ROWS, ws.max_row or 1)
    for r in range(1, scan + 1):
        code_c = 0
        for c in range(1, (ws.max_column or 1) + 1):
            if _match(_norm(ws.cell(row=r, column=c).value), ALIAS_CODE):
                code_c = c
                break
        if not code_c:
            continue
        cols = {"code": code_c, "name": 0, "supplier": 0,
                "demand": 0, "remain": 0, "header_row": r}
        supplier_fallback = 0
        for c in range(1, (ws.max_column or 1) + 1):
            t = _norm(ws.cell(row=r, column=c).value)
            if not t:
                continue
            if cols["name"] == 0 and _match(t, ALIAS_NAME):
                cols["name"] = c
            if cols["demand"] == 0 and _match(t, ALIAS_DEMAND):
                cols["demand"] = c
            if cols["remain"] == 0 and _match(t, ALIAS_REMAIN):
                cols["remain"] = c
            if "供应商" in t:
                if "信息" in t and cols["supplier"] == 0:
                    cols["supplier"] = c
                elif ("代码" not in t and "编号" not in t
                      and "代号" not in t and supplier_fallback == 0):
                    supplier_fallback = c
        if cols["supplier"] == 0:
            cols["supplier"] = supplier_fallback
        # 关键列齐了才算成功; 名称/供应商缺失可空着(不致命)
        if cols["code"] and cols["demand"] and cols["remain"]:
            return cols
    return None


# 数据形态画像:编码/名称/供应商/需求/剩余未收。code 与两数值列为必需(与 locate 一致)。
_SHAPE_PROFILE = [
    ("code", shape_detect.CODE, True),
    ("name", shape_detect.TEXT, False),
    ("supplier", shape_detect.TEXT, False),
    ("demand", shape_detect.NUMBER, True),
    ("remain", shape_detect.NUMBER, True),
]


def _locate_by_shape(ws, log=None):
    """表头文字识别失败时,按数据形态兜底猜列。返回与 locate_columns 同构的 dict 或 None。

    命中后由上层(UI)交用户核对再生成;命令行/自动路径也会在日志醒目提示。"""
    hr, col, _conf = shape_detect.detect_by_shape(
        ws, _SHAPE_PROFILE, scan_rows=HEADER_SCAN_ROWS, log=log)
    if not hr:
        return None
    return {"code": col["code"], "name": col.get("name", 0),
            "supplier": col.get("supplier", 0), "demand": col["demand"],
            "remain": col["remain"], "header_row": hr}

def detect_batch(path):
    """从表内A1标题优先识别批次号, 失败则从文件名识别, 再失败返回空串"""
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        a1 = _first_ws(wb)['A1'].value or ""
        wb.close()
        m = RE_TITLE.search(str(a1))
        if m:
            return m.group(1)
    except Exception:
        pass
    m = RE_NAME.search(os.path.basename(path))
    return m.group(1) if m else ""

def find_plan_files(folder):
    """返回文件夹内所有送货计划表路径(排除临时文件), 按修改时间倒序"""
    files = [f for f in glob.glob(os.path.join(folder, "*送货计划*.xlsx"))
             if not os.path.basename(f).startswith("~$")]
    return sorted(files, key=os.path.getmtime, reverse=True)

def extract_unreceived(path, log=None):
    """提取未收货物料: [编码, 名称, 供应商, 需求数, 剩余未收数].
       规则:
       1) 列位置按表头文字灵活识别(locate_columns), 失败才回退固定列, 这样
          即便收到的表列顺序/多出几列, 也不会错位;
       2) 表头行之后才是数据(不再写死从第3行开始);
       3) 跳过被筛选隐藏的行(尊重用户在 Excel 里的筛选);
       4) "剩余未收数"为 #N/A / #REF! 等错误值 / 空 / 0 -> 当作已收, 排除;
          只保留为非零数值的行。"""
    import numbers
    # 剩余未收数常是公式; 若该表未被 Excel 刷新过, data_only 读出 None,
    # 会把整列误当"已收"静默排除(全部物料显示已收)。读表前先醒目告警。
    if log:
        warn_if_uncached(path, log, what="剩余未收数")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = _pick_data_ws(wb, log=log)
    cols = locate_columns(ws)
    if cols is None:                                # 表头文字识别失败: 先试数据形态兜底
        cols = _locate_by_shape(ws, log=log)
    if cols is None:                                # 都失败才回退老的固定列
        cols = {"code": COL_CODE, "name": COL_NAME, "supplier": COL_SUPPLIER,
                "demand": COL_DEMAND, "remain": COL_REMAIN, "header_row": 2}
    c_code, c_name = cols["code"], cols["name"]
    c_supp, c_dem, c_rem = cols["supplier"], cols["demand"], cols["remain"]
    start = cols["header_row"] + 1
    rows = []
    pending = 0                                     # 剩余未收数读出 None 的行数(疑似公式未刷新)
    for r in range(start, ws.max_row + 1):
        dim = ws.row_dimensions.get(r)
        if dim is not None and dim.hidden:          # 被筛选隐藏 -> 跳过
            continue
        code = ws.cell(row=r, column=c_code).value
        if code is None:
            continue
        remain = ws.cell(row=r, column=c_rem).value
        if isinstance(remain, bool):
            continue
        # remain 为 None: 公式未刷新读不到值, 不能当 0/已收静默排除, 单独计数提示
        if remain is None:
            pending += 1
            continue
        # #N/A 等错误(字符串) / 0 -> 当作已收, 排除; 只认非零数值
        if not isinstance(remain, numbers.Number) or remain == 0:
            continue
        rows.append([str(code),
                     ws.cell(row=r, column=c_name).value if c_name else None,
                     ws.cell(row=r, column=c_supp).value if c_supp else None,
                     ws.cell(row=r, column=c_dem).value if c_dem else None,
                     remain])
    # 有剩余未收数读不到值的行 -> 提示人工确认, 而非当已收吞掉
    if pending and log:
        log("⚠ 《%s》有 %d 行的剩余未收数读不到值(可能公式未刷新), 未计入未收料, 请核对。"
            % (os.path.basename(path), pending))
    return rows
def _style(cell, bold=False, fill=False, align=CENTER):
    cell.border = BORDER
    cell.alignment = align
    cell.font = FONT_B if bold else FONT
    if fill:
        cell.fill = BLUE_FILL

def _write_batch(ws, c0, batch_no, materials, total, remark, top_label):
    diff = len(materials)
    arrived = total - diff
    # 第1、2行: 截止标签 / 批次号 (跨C:E, 白底)
    for row, val in [(1, top_label), (2, batch_no)]:
        ws.cell(row=row, column=c0, value=val)
        ws.merge_cells(start_row=row, start_column=c0, end_row=row, end_column=c0 + 2)
        for cc in range(c0, c0 + 3):
            _style(ws.cell(row=row, column=cc))
    # 第3-5行: 标签(蓝底,跨2列) + 数值(白底)
    for row, label, val in [(3, "主料总共类", total), (4, "到货数量", arrived), (5, "差异", diff)]:
        ws.cell(row=row, column=c0, value=label)
        ws.merge_cells(start_row=row, start_column=c0, end_row=row, end_column=c0 + 1)
        for cc in range(c0, c0 + 2):
            _style(ws.cell(row=row, column=cc), fill=True)   # 标签蓝底
        ws.cell(row=row, column=c0 + 2, value=val)
        _style(ws.cell(row=row, column=c0 + 2))              # 数值白底
    # 第6行: 备注(有则写, 跨7列, 左对齐, 白底)
    if remark:
        ws.cell(row=6, column=c0, value="备注: " + str(remark))
        ws.merge_cells(start_row=6, start_column=c0, end_row=6, end_column=c0 + 6)
        for cc in range(c0, c0 + 7):
            _style(ws.cell(row=6, column=cc), align=LEFT)
    # 第7行: 表头(蓝底加粗) —— 新增"剩余未收数"列
    for i, h in enumerate(["序号", "物料编码", "物料名称", "供应商信息",
                           "需求数", "剩余未收数", "备注"]):
        _style(ws.cell(row=7, column=c0 + i, value=h), bold=True, fill=True)
    # 数据行(白底)
    for i, (code, name, supp, dem, remain) in enumerate(materials):
        r = DATA_START_ROW + i
        for j, v in enumerate([i + 1, code, name, supp, dem, remain, None]):
            _style(ws.cell(row=r, column=c0 + j, value=v))
    for i, w in enumerate(COL_WIDTHS):
        ws.column_dimensions[get_column_letter(c0 + i)].width = w
    return diff, arrived

def build_workbook(batches, top_label, out_path):
    """batches: [{batch_no, materials, total, remark}, ...]"""
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Sheet1"
    results = []
    col = FIRST_COL
    for b in batches:
        diff, arrived = _write_batch(ws, col, b["batch_no"], b["materials"],
                                     b["total"], b.get("remark", ""), top_label)
        results.append((b["batch_no"], diff, arrived, b["total"]))
        col += BATCH_STRIDE
    # 与 purchase/delivery 一致: 目标被 Excel 占用时给出友好提示
    try:
        wb.save(out_path)
    except PermissionError:
        raise PermissionError("无法保存 %s —— 请先在 Excel 里关闭该文件后重试" % out_path)
    return results


# ---------------- 统一入口（与其它三功能同构：接受统一 out_dir）----------------
def build_batches(rows_data, top_label, log=None):
    """把界面收集的每行数据整理成 build_workbook 需要的批次列表，并顺带
    读取每个送货计划表的未收料明细。纯逻辑，不依赖界面。
    rows_data: [{path, batch_no, total, remark, include(bool)}, ...]
    返回 (batches, mem)：batches 供 build_workbook；mem 为批次记忆(供落盘)。
    """
    batches, mem = [], {}
    for row in rows_data:
        if not row.get("include", True):
            continue
        materials = extract_unreceived(row["path"], log=log)
        bn = row.get("batch_no") or detect_batch(row["path"])
        # 界面留空时 total 可能为 None/""/带小数或千分位的文本; int("566.0")、
        # int("5,66") 都会抛 ValueError, 统一 float() 兜一层再取整, 失败回退默认值
        tv = row.get("total", DEFAULT_TOTAL)
        try:
            total = DEFAULT_TOTAL if tv in (None, "") else int(float(str(tv).replace(",", "")))
        except (ValueError, TypeError):
            total = DEFAULT_TOTAL
        remark = row.get("remark", "")
        batches.append({"batch_no": bn, "materials": materials,
                        "total": total, "remark": remark})
        if bn:
            mem[bn] = {"total": total, "remark": remark}
    return batches, mem


def run(rows_data, top_label=None, out_dir=None, log=None):
    """到料明细统一入口。
    rows_data : 见 build_batches。
    out_dir   : 输出目录；不传则用统一 paths 系统（文档下统一文件夹）。
    返回 {"out_file", "out_dir", "results"}。
    """
    log = log or (lambda *a, **k: None)
    st = _settings.get_settings()
    if top_label is None:
        top_label = st.arrival.get("top_label", DEFAULT_TOP_LABEL)
    if out_dir is None:
        out_dir = _paths.resolve_output_dir("arrival", **st.output_kwargs())
    log("整理批次数据（共 %d 个计划表）..." % len(rows_data))
    batches, mem = build_batches(rows_data, top_label, log=log)
    fname = "%s每日主料到料明细.xlsx" % beijing_date()
    out_file = os.path.join(out_dir, fname)
    results = build_workbook(batches, top_label, out_file)
    for bn, diff, arrived, total in results:
        log("  · 批次 %s：未收料 %d 类，到货 %d，主料 %d 类" % (bn, diff, arrived, total))
    # 更新批次记忆到全局设置
    st.arrival["top_label"] = top_label
    st.arrival.setdefault("batches", {}).update(mem)
    if batches:
        st.arrival["last_total"] = batches[-1]["total"]
    st.save()
    log("已保存：%s" % out_file)
    return {"out_file": out_file, "out_dir": out_dir, "results": results}
