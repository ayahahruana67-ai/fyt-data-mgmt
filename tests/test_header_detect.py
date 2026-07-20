# -*- coding: utf-8 -*-
"""共享表头识别引擎 header_detect 的直接单元测试(合成簿,可移植)。

钉死引擎级不变量:先精确后包含、require 全含才算表头、exclude_contains
只在包含匹配阶段生效、dict 顺序即列优先级。delivery/purchase 的封装测试
另见 test_delivery_headers / test_integration。
"""
import os
import tempfile
import unittest
import warnings

import openpyxl

from core import header_detect as H

warnings.filterwarnings("ignore", message="Workbook contains no default style")

KEYS = {
    "code": ["零部件代码", "编码"],
    "name": ["名称", "品名"],
    "qty":  ["数量"],
    "sup_name": ["供应商名称", "供应商"],
}


class _Tmp(unittest.TestCase):
    def setUp(self):
        self._tmp = tempfile.mkdtemp(prefix="fyt_hd_")

    def tearDown(self):
        import shutil
        shutil.rmtree(self._tmp, ignore_errors=True)

    def cols(self, header, **kw):
        p = os.path.join(self._tmp, "t.xlsx")
        wb = openpyxl.Workbook(); ws = wb.active
        ws.append(header); ws.append(["x"] * len(header)); wb.save(p)
        wb2 = openpyxl.load_workbook(p)
        hr, cols = H.detect_layout(wb2[wb2.sheetnames[0]], KEYS, **kw)
        wb2.close()
        return hr, cols


class TestEngine(_Tmp):
    def test_exact_before_contains(self):
        # "编码"(含匹配 code)不应抢占精确的"零部件代码";"数量"精确命中 qty
        hr, c = self.cols(["零部件代码", "名称", "数量"], require=("code",))
        self.assertEqual(hr, 1)
        self.assertEqual(c["code"], 1)

    def test_require_all_present(self):
        # require 含 qty 但表里无数量列 -> 不认为是表头
        hr, c = self.cols(["零部件代码", "名称"], require=("code", "qty"))
        self.assertIsNone(hr)

    def test_require_met(self):
        hr, c = self.cols(["零部件代码", "名称", "数量"], require=("code", "qty"))
        self.assertIsNotNone(hr)

    def test_exclude_contains_blocks_interference(self):
        # 「委外供应商属性」含"供应商"但被 exclude_contains 挡在包含匹配之外
        hr, c = self.cols(["零部件代码", "委外供应商属性", "供应商名称"],
                          require=("code",),
                          exclude_contains={"sup_name": ["属性"]})
        self.assertEqual(c["sup_name"], 3)      # 指向真正的供应商名称列

    def test_exclude_does_not_affect_exact(self):
        # 精确匹配不受 exclude_contains 影响:列名恰为"供应商名称"仍命中
        hr, c = self.cols(["零部件代码", "供应商名称"],
                          require=("code",),
                          exclude_contains={"sup_name": ["供应商"]})
        self.assertEqual(c["sup_name"], 2)

    def test_best_row_most_roles(self):
        # 第2行命中更多角色 -> 选第2行为表头
        p = os.path.join(self._tmp, "m.xlsx")
        wb = openpyxl.Workbook(); ws = wb.active
        ws.append(["零部件代码", "杂", "乱"])              # 行1:仅 code
        ws.append(["零部件代码", "名称", "数量"])          # 行2:code+name+qty
        wb.save(p)
        wb2 = openpyxl.load_workbook(p)
        hr, c = H.detect_layout(wb2[wb2.sheetnames[0]], KEYS, require=("code",))
        wb2.close()
        self.assertEqual(hr, 2)

    def test_no_header_returns_none(self):
        hr, c = self.cols(["甲", "乙", "丙"], require=("code",))
        self.assertIsNone(hr)
        self.assertEqual(c, {})

    def test_log_reports_unmatched_columns(self):
        # 有列但无别名 -> 应通过 log 上报,避免静默漏读(如当初的"下阶物料")
        logs = []
        self.cols(["零部件代码", "名称", "数量", "交期", "库位"],
                  require=("code",), log=logs.append)
        joined = " ".join(logs)
        self.assertIn("交期", joined)
        self.assertIn("库位", joined)

    def test_log_silent_when_all_matched(self):
        # 全列认领时不产生未认领日志(不刷屏)
        logs = []
        self.cols(["零部件代码", "名称", "数量"], require=("code",), log=logs.append)
        self.assertEqual([l for l in logs if "未认领" in l], [])

    def test_log_none_is_default_noop(self):
        # 不传 log 时行为与原先完全一致(纯增量特性)
        hr, c = self.cols(["零部件代码", "名称", "数量", "交期"], require=("code",))
        self.assertIsNotNone(hr)
        self.assertNotIn("交期", c)          # 交期无别名,不应被认领

    def test_inner_whitespace_collapsed_for_exact(self):
        # 表头含手动换行/内部空格/全角空格时,折叠后仍能精确命中关键词,
        # 不因空白退化到包含匹配被别列抢占。这是识别率的常见真实缺口。
        hr, c = self.cols(["零部件\n代码", "供应商 名称", "数　量"],
                          require=("code",))
        self.assertEqual(hr, 1)
        self.assertEqual(c["code"], 1)          # "零部件\n代码" -> "零部件代码" 精确命中
        self.assertEqual(c["sup_name"], 2)      # "供应商 名称" -> "供应商名称" 精确命中
        self.assertEqual(c["qty"], 3)           # 全角空格也被折叠

    def test_whitespace_header_still_excluded(self):
        # 折叠空白后 exclude_contains 仍生效:"委外 供应商\n属性" -> 含"属性"被挡
        hr, c = self.cols(["零部件代码", "委外 供应商\n属性", "供应商名称"],
                          require=("code",),
                          exclude_contains={"sup_name": ["属性"]})
        self.assertEqual(c["sup_name"], 3)


if __name__ == "__main__":
    unittest.main()
