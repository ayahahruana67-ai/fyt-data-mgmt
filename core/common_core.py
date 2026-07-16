# -*- coding: utf-8 -*-
"""
公共核心模块 —— 考勤填报 与 工时对账 两功能共用
=================================================
把两个功能原先各自实现、且细节不一致的工具函数统一到这里，避免同名不同义：
  · 姓名规范化：统一为“去掉所有空格”（原填报只去首尾，会导致“张 三”对不上）
  · 日期、时间、数字解析：各一份权威实现
  · read_sheets：统一的多格式读取（.xlsx/.xlsm/.xls）
  · 常量：标准工时、容差、假/休标记
  · Options：高级选项载体（供用户手动干预识别）
  · 输出路径：两功能统一「源文件目录/output/名字_后缀_时间戳.ext」

兼容 Windows 7 + Python 3.8。
"""
import os
import re
import datetime

import openpyxl

# ---------------- 统一常量 ----------------
# 标准工作时长（小时）：实际工时超出此值的部分记为加班，不足记 0。
STANDARD_WORKDAY_HOURS = 9.0
# 数值容差（工时以 0.5 为最小单位，容差取 0.01 判断相等）
TOL = 0.01
# 视为“非工时”的标记（假/休/空白等），不参与逐日对比
SKIP_MARKS = {"假", "休", "旷", "缺", "调休", "病假", "事假", "年假", ""}


# ---------------- 列角色定义（供可视化列映射界面使用） ----------------
# 每种“文件类型”需要指定哪些列角色。key=角色内部名，value=(中文名, 是否必填)
ROLE_DEFS = {
    "att_source": [("name", "姓名", True), ("date", "日期", True),
                   ("on", "上班1打卡时间", True), ("off", "下班1打卡时间", True)],
    "att_target": [("name", "姓名", True), ("date", "日期", True),
                   ("sys_on", "上班时间(系统)", False), ("act_on", "上班时间(实际)", False),
                   ("sys_off", "下班时间(系统)", False), ("act_off", "下班时间(实际)", False),
                   ("rest", "休息时间", False), ("work", "实际工作时间", False),
                   ("ot", "加班", False)],
    "rec_source": [("name", "姓名", True), ("date", "日期", True),
                   ("work", "实际工作时间", True)],
    "rec_zong": [("name", "姓名", True), ("comp", "所属劳务公司", False),
                 ("work", "出勤工时", False), ("check", "对账时间", False)],
    "rec_labor": [("name", "姓名", True), ("total", "合计/出勤工时列", False)],
}
KIND_TITLES = {
    "att_source": "填报·系统数据表", "att_target": "填报·待填考勤表",
    "rec_source": "对账·数据来源", "rec_zong": "对账·待对总表",
    "rec_labor": "对账·劳务对账单",
}


# ---------------- 高级选项 ----------------
class Options:
    """高级选项载体：让用户在程序识别有误时手动干预。

    workday_hours : 每日标准工时（小时），加班基准。
    overtime      : 是否计算加班列。
    conflict      : 同一(姓名,日期)在多文件重复时的策略 last/first/warn。
    header_row    : 全局手动表头行(1-based)；None=自动。（per-file 映射优先）
    sheet_name    : 全局手动工作表名；None=自动/全部。（per-file 映射优先）
    tolerance     : 对账工时比对容差(小时)。
    data_start    : 手动数据起始行(1-based)；None=表头下一行。（per-file 映射优先）
    skip_extra    : 追加的“非工时”标记词集合（与内置 SKIP_MARKS 合并）。
    columns       : per-file 列映射 {文件名basename: {"sheet":名或None, "header":行1based或None,
                    "data_start":行或None, "roles":{角色:列0based}}}。手动映射优先于自动识别。
    auto_actual   : 是否自动按半小时算“实际上/下班时间”（上班进位、下班退位）。默认 True。
    night_shift   : 是否启用两班制夜班识别（跨零点 +24 修正）。默认 True。
    night_start_hour   : 实际上班钟点 ≥ 此值判为夜班。默认 17.0。
    night_workday_hours: 夜班标准工时（加班基准）。默认 11.0。
    night_max_hours    : 夜班合理工时上限，超过判异常（防漏打卡）。默认 16.0。
    """
    def __init__(self, workday_hours=STANDARD_WORKDAY_HOURS, overtime=True,
                 conflict="last", header_row=None, sheet_name=None, tolerance=TOL,
                 data_start=None, skip_extra=None, columns=None,
                 auto_actual=True, night_shift=True, night_start_hour=17.0,
                 night_workday_hours=11.0, night_max_hours=16.0):
        self.workday_hours = float(workday_hours)
        self.overtime = bool(overtime)
        # 自动按半小时算“实际上/下班时间”：上班进位、下班退位，再据此算实际工时。
        self.auto_actual = bool(auto_actual)
        # 两班制：按实际上班钟点区分白/夜班；夜班跨零点自动 +24 修正。
        self.night_shift = bool(night_shift)              # 是否启用夜班识别
        self.night_start_hour = float(night_start_hour)   # 上班打卡≥此钟点 → 夜班
        self.night_workday_hours = float(night_workday_hours)  # 夜班标准工时(加班基准)
        self.night_max_hours = float(night_max_hours)     # 夜班合理工时上限(超则判异常)
        self.conflict = conflict if conflict in ("last", "first", "warn") else "last"
        self.header_row = header_row
        self.sheet_name = (sheet_name or None)
        self.tolerance = float(tolerance)
        self.data_start = data_start
        self.skip_extra = set(skip_extra) if skip_extra else set()
        self.columns = columns if columns else {}

    def skip_set(self):
        """内置 + 自定义 的“非工时”标记词集合。"""
        return SKIP_MARKS | self.skip_extra

    def file_map(self, path):
        """取某文件的 per-file 列映射；无则 None。path 可为完整路径或 basename。"""
        if not self.columns:
            return None
        return self.columns.get(os.path.basename(path)) or self.columns.get(path)

    def resolve_sheet(self, path):
        """该文件应处理的工作表名：per-file > 全局 > None(自动)。"""
        fm = self.file_map(path)
        if fm and fm.get("sheet"):
            return fm["sheet"]
        return self.sheet_name

    def resolve_header(self, path):
        """该文件表头行(1-based)：per-file > 全局 > None。"""
        fm = self.file_map(path)
        if fm and fm.get("header"):
            return fm["header"]
        return self.header_row

    def resolve_data_start(self, path):
        """该文件数据起始行(1-based)：per-file > 全局 > None。"""
        fm = self.file_map(path)
        if fm and fm.get("data_start"):
            return fm["data_start"]
        return self.data_start

    def resolve_roles(self, path):
        """该文件手动列映射 {角色:列0based}；无则 {}。"""
        fm = self.file_map(path)
        return dict(fm["roles"]) if (fm and fm.get("roles")) else {}

    def summary(self):
        """一行文字，用于日志追溯本次采用的选项。"""
        cn = {"last": "后者覆盖", "first": "先者优先", "warn": "不覆盖仅提示"}
        parts = ["标准工时=%g" % self.workday_hours,
                 "加班=%s" % ("算" if self.overtime else "不算"),
                 "重复=%s" % cn.get(self.conflict, self.conflict),
                 "容差=%g" % self.tolerance]
        if self.header_row:
            parts.append("表头行=%d" % self.header_row)
        if self.sheet_name:
            parts.append("工作表=%s" % self.sheet_name)
        if self.skip_extra:
            parts.append("额外假休标记=%s" % "/".join(sorted(self.skip_extra)))
        parts.append("实际时间=%s" % ("自动半小时进退位" if self.auto_actual else "不自动"))
        if self.night_shift:
            parts.append("夜班=启用(≥%g点/标准%gh/上限%gh)"
                         % (self.night_start_hour, self.night_workday_hours, self.night_max_hours))
        else:
            parts.append("夜班=不识别")
        if self.columns:
            parts.append("列映射=%d个文件" % len(self.columns))
        return "；".join(parts)


DEFAULTS = Options()


# ---------------- 统一解析工具 ----------------
def norm_name(v):
    """姓名规范化：去掉所有空格（含中间空格）。两功能一致，避免“张 三”对不上。"""
    if v is None:
        return ""
    return re.sub(r"\s+", "", str(v))


def norm_date(v):
    """把各种日期形式统一成 (year, month, day) 元组，无法解析返回 None。"""
    if v is None:
        return None
    if isinstance(v, (datetime.datetime, datetime.date)):
        return (v.year, v.month, v.day)
    s = str(v).strip()
    if not s or s == "-":
        return None
    digits = "".join(ch for ch in s if ch.isdigit())
    if len(digits) == 8:                       # 20260501
        try:
            return (int(digits[0:4]), int(digits[4:6]), int(digits[6:8]))
        except ValueError:
            pass
    for sep in ("-", "/", "."):                # 2026-05-01 / 2026/5/1
        if sep in s:
            parts = s.split(sep)
            if len(parts) == 3:
                try:
                    return (int(parts[0]), int(parts[1]), int(parts[2]))
                except ValueError:
                    pass
    return None


def day_of(v):
    """把各种日期表示转成“当月第几天”(1~31)。识别失败返回 None。"""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.day
    if isinstance(v, (int, float)):
        n = int(v)
        return n if 1 <= n <= 31 else None
    s = str(v).strip()
    if re.fullmatch(r"\d{8}", s):              # 20260501
        return int(s[6:8])
    m = re.search(r"(\d{1,2})\s*日?\s*$", s)   # 结尾数字或“5日”
    if m:
        n = int(m.group(1))
        if 1 <= n <= 31:
            return n
    return None


def parse_time(v):
    """把单元格值解析成 datetime.time；无效返回 None。"""
    if v is None:
        return None
    if isinstance(v, datetime.time):
        return v
    if isinstance(v, datetime.datetime):
        return v.time()
    s = str(v).strip()
    if not s or s in ("-", "—"):
        return None
    if ":" in s:
        parts = s.split(":")
        try:
            h = int(parts[0]); m = int(parts[1])
            sec = int(parts[2]) if len(parts) > 2 else 0
            return datetime.time(h, m, sec)
        except (ValueError, IndexError):
            return None
    return None


def to_hours(t):
    """datetime.time -> 小时数(float)。None -> None。"""
    if t is None:
        return None
    return t.hour + t.minute / 60.0 + t.second / 3600.0


def round_half_hour(t, mode):
    """按半小时把 datetime.time 取整到最近的整点/半点。

    mode="up"  进位（用于上班）：向上取到 :00 或 :30。7:56→8:00，7:31→8:00，7:30→7:30。
    mode="down"退位（用于下班）：向下取到 :00 或 :30。8:13→8:00，8:24→8:00，8:30→8:30。
    恰好落在整点/半点则不变。t 为 None 返回 None。
    """
    import math
    if t is None:
        return None
    total = t.hour * 3600 + t.minute * 60 + t.second      # 当日秒数
    step = 1800                                            # 半小时
    if mode == "up":
        secs = int(math.ceil(total / float(step))) * step
    else:                                                  # down
        secs = (total // step) * step
    secs = max(0, min(secs, 86400 - step))                # 兜底钳制（考勤为日间，进位不会到 24:00）
    return datetime.time(secs // 3600, (secs % 3600) // 60)


def fmt_time(t):
    """datetime.time -> 'HH:MM' 字符串；None -> ''。"""
    if t is None:
        return ""
    return "%02d:%02d" % (t.hour, t.minute)


def parse_rest(v):
    """休息时间解析为小时数(float)。空/无效 -> 0。"""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s or s in ("-", "—"):
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def to_num(v, skip=None):
    """把单元格值转成 float；无法转（含假/休/空白）返回 None。
    skip: 可选“非工时”标记词集合，默认用内置 SKIP_MARKS。"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    marks = skip if skip is not None else SKIP_MARKS
    if s == "" or s in marks:
        return None
    try:
        return float(s)
    except ValueError:
        return None


# ---------------- 统一多格式读取 ----------------
def read_sheets(path):
    """读取任意工作簿为 [(sheet_name, rows)]，rows 为二维列表。
    支持 .xlsx / .xlsm / .xls（.xls 需 xlrd==1.2.0）。两功能共用。"""
    ext = os.path.splitext(path)[1].lower()
    result = []
    if ext in (".xlsx", ".xlsm"):
        wb = openpyxl.load_workbook(path, data_only=True)
        try:
            for ws in wb.worksheets:
                result.append((ws.title,
                               [list(row) for row in ws.iter_rows(values_only=True)]))
        finally:
            wb.close()
    elif ext == ".xls":
        import xlrd  # 打包时通过 xlrd==1.2.0 支持 .xls
        book = xlrd.open_workbook(path)
        for sh in book.sheets():
            rows = []
            for r in range(sh.nrows):
                row = []
                for c in range(sh.ncols):
                    cell = sh.cell(r, c)
                    val = cell.value
                    if cell.ctype == 3:        # xlrd 类型3=日期
                        try:
                            y, mo, d, h, mi, s = xlrd.xldate_as_tuple(val, book.datemode)
                            val = datetime.datetime(y, mo, d, h, mi, s)
                        except Exception:
                            pass
                    row.append(val)
                rows.append(row)
            result.append((sh.name, rows))
    else:
        raise ValueError("不支持的文件类型：%s" % ext)
    return result


# ---------------- 统一输出路径 ----------------
def make_out_dir(src_path):
    """在源文件所在目录建 output 子文件夹并返回其路径（两功能统一）。"""
    base_dir = os.path.dirname(os.path.abspath(src_path))
    out_dir = os.path.join(base_dir, "output")
    if not os.path.isdir(out_dir):
        os.makedirs(out_dir)
    return out_dir


def timestamp():
    """当前时间戳字符串 YYYYMMDD_HHMM，用于输出文件名。"""
    return datetime.datetime.now().strftime("%Y%m%d_%H%M")


def out_path(out_dir, base_name, suffix, ext=".xlsx", ts=None):
    """拼输出路径：out_dir/base_name_suffix_时间戳.ext（两功能统一命名）。
    base_name 传源文件名(不含扩展名)；ts 不传则自动取当前时间。"""
    if ts is None:
        ts = timestamp()
    fname = "%s%s_%s%s" % (base_name, suffix, ts, ext)
    return os.path.join(out_dir, fname)


# ---------------- 供“列映射”界面读取表头预览 ----------------
def sheet_names(path):
    """只取工作表名列表（供界面下拉）。"""
    return [name for name, _ in read_sheets(path)]


def preview_rows(path, sheet=None, limit=8):
    """取某工作表前 limit 行（含表头，二维列表）。sheet=None 取第一个。
    返回 (sheet_name, rows)。供列映射界面显示表头与样例数据。"""
    sheets = read_sheets(path)
    if not sheets:
        return None, []
    chosen = None
    if sheet:
        for name, rows in sheets:
            if name == sheet:
                chosen = (name, rows); break
    if chosen is None:
        chosen = sheets[0]
    name, rows = chosen
    return name, [list(r) for r in rows[:limit]]


def cell_text(v):
    """单元格值转成简短显示文本（供界面）。"""
    if v is None:
        return ""
    if isinstance(v, datetime.datetime):
        return v.strftime("%Y-%m-%d %H:%M").replace(" 00:00", "")
    if isinstance(v, datetime.date):
        return v.strftime("%Y-%m-%d")
    s = str(v)
    return s if len(s) <= 18 else s[:17] + "…"
