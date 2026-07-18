# -*- coding: utf-8 -*-
"""
送货计划表制作 —— 核心逻辑
==========================
以"物料清单"为主表逐行生成送货计划，按物料号从"物料明细表(含供应商)"查供应商
代码与名称；KD/SUB 列按用户选择统一填写；CASE/CASE托数/班组 可从一张已做好的
往期送货计划(参考表)按物料编码带出；到货/收货等跟单列留空供后续人工填写。

输入（顺序任意，程序自动辨识）：
  · 物料清单：含 物料号 + 数量（可再含中/英文描述）——决定输出的行与需求数；
  · 供应商明细：含 零部件代码 + 供应商代码 + 供应商名称——供按编码查供应商；
  · 参考送货计划（可选）：往期做好的送货计划，按物料编码带出 CASE/CASE托数/班组。

输出 16 列送货计划，样式与客户新模板一致（标题行合并留空、微软雅黑、全边框，
"剩余未收数"= 实际收货数 - 需求数 的公式）。表头行与列位置自动识别，不写死列号。
兼容 Windows 7 + Python 3.8 + PySide2。
"""
import os

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import paths as _paths
from . import settings as settings_mod
from . import header_detect
from . import shape_detect

# 表头关键词 -> 角色。识别时先精确匹配，再包含匹配；每列只归一个角色。
HEADER_KEYS = {
    # "下阶物料描述" 放前不影响(精确匹配优先);"下阶物料" 是 SAP KD 清单/供应商表的编码列
    "code":  ["物料号", "物料编码", "零部件代码", "物料编号", "下阶物料", "零件号", "料号", "编码"],
    "cname": ["物料中文描述", "物料英文描述", "下阶物料描述", "物料名称", "零部件名称", "中文描述", "名称", "品名"],
    "ename": ["物料英文描述", "英文描述", "英文名称"],
    "qty":   ["需求数", "需求数量", "计划数量", "数量"],
    "sup_code": ["供应商代码", "供应商编码", "供方代码"],
    "sup_name": ["供应商名称", "供应商信息", "供方名称", "供应商"],
    "attr":  ["属性", "KD/SUB", "KD/SUB属性"],
}

# 输出固定 16 列（顺序即样本顺序）
OUT_HEADERS = ["序号", "物料编码", "物料名称", "供应商代码", "供应商信息", "KD/SUB",
               "需求数", "计划到货日期", "实际收货数", "实际收货日期", "第二次到货日期",
               "剩余未收数", "CASE", "CASE托数", "班组", "备注"]


def norm_code(v):
    """物料号归一：转字符串去空格。数值型编码去掉尾随的 .0。"""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def cell_text(v):
    """单元格文本化：None->''，浮点整数去 .0，其余原样 str。"""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


# ---------------------------------------------------------------------------
# 表头/列自动识别
# ---------------------------------------------------------------------------
# 包含匹配时,"委外供应商属性"这类标志列含"供应商"子串但实为属性列,
# 不得当作供应商代码/名称;精确匹配不受影响。守卫随共享引擎走(见 header_detect)。
_EXCLUDE_CONTAINS = {"sup_code": ["属性"], "sup_name": ["属性"]}


def detect_layout(ws, scan_rows=12, log=None):
    """在前若干行里找表头行并映射列。返回 (header_row, {角色:列号})。

    选"能命中最多角色"的行为表头行；要求至少含 code 列，否则视为未识别(返回 None,{})。
    薄封装:算法在 header_detect.detect_layout,此处只传本功能的常量。
    """
    return header_detect.detect_layout(
        ws, HEADER_KEYS, require=("code",), scan_rows=scan_rows,
        exclude_contains=_EXCLUDE_CONTAINS, log=log)


# 数据形态画像(供 header_detect 失败时兜底):角色顺序≈典型列序,末位 bool 为必需。
_SHAPE_PROFILE = [
    ("code", shape_detect.CODE, True),
    ("cname", shape_detect.TEXT, False),
    ("qty", shape_detect.NUMBER, False),
    ("sup_code", shape_detect.CODE, False),
    ("sup_name", shape_detect.TEXT, False),
]


def detect_layout_or_shape(ws, scan_rows=12, log=None):
    """先按表头文字识别;失败再按数据形态兜底。

    返回 (header_row, col_map, source):source 为 "header"(表头识别)或
    "shape"(形态兜底,需用户确认)或 None(都失败)。兜底命中的映射**不应静默落盘**,
    由调用方交用户核对。"""
    hr, col = detect_layout(ws, scan_rows=scan_rows, log=log)
    if hr:
        return hr, col, "header"
    hr2, col2, _conf = shape_detect.detect_by_shape(
        ws, _SHAPE_PROFILE, scan_rows=scan_rows, log=log)
    if hr2:
        return hr2, col2, "shape"
    return None, {}, None


def list_sheets(path):
    """列出工作簿的子表名(供界面下拉选择)。读失败/非 xlsx 返回 []。"""
    ext = os.path.splitext(path)[1].lower()
    if ext not in (".xlsx", ".xlsm"):
        return []                        # openpyxl 不读 .xls;交由默认(第一表)处理
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
        try:
            return list(wb.sheetnames)
        finally:
            wb.close()
    except Exception:
        return []


def load_sheet(path, sheet=None, log=None):
    """读取一张表：自动识别表头与列。返回 (rows, layout)。

    rows: [{r, code, cname, ename, qty, sup_code, sup_name, attr}]，按角色缺省为 None。
    已过滤空编码行与合计/小计行。
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
        header_row, col = detect_layout(ws, log=log)
        if not header_row:
            raise ValueError("未能在 %s / %s 中识别表头（需含“物料号/编码”列）"
                             % (os.path.basename(path), ws.title))
        rows = []
        for r in range(header_row + 1, (ws.max_row or header_row) + 1):
            code = ws.cell(r, col["code"]).value
            if code is None or norm_code(code) == "":
                continue
            cn = ws.cell(r, col["cname"]).value if "cname" in col else None
            if isinstance(cn, str) and ("合计" in cn or "小计" in cn or "总计" in cn):
                continue
            rows.append({
                "r": r, "code": code, "cname": cn,
                "ename": ws.cell(r, col["ename"]).value if "ename" in col else None,
                "qty": ws.cell(r, col["qty"]).value if "qty" in col else None,
                "sup_code": ws.cell(r, col["sup_code"]).value if "sup_code" in col else None,
                "sup_name": ws.cell(r, col["sup_name"]).value if "sup_name" in col else None,
                "attr": ws.cell(r, col["attr"]).value if "attr" in col else None,
            })
        return rows, {"sheet": ws.title, "header_row": header_row, "col": col}
    finally:
        wb.close()


def analyze(path, sheet=None, log=None):
    """选择即扫描:只读地预检一个文件的表头识别结果,不生成任何文件。

    供 UI 在用户选文件后立刻反馈"能否认出各列 / 是靠形态兜底(需核对)",
    以便在点"生成"前就发现列错位。返回 dict:
      ok           - 是否识别成功(拿到必需列)
      sheet        - 命中的子表名
      header_row   - 表头所在行(1-based)
      roles        - {角色: 列号}
      source       - "header"(表头文字识别) / "shape"(形态兜底,需人工核对) / None
      n_rows       - 表头之后的数据行数(粗计,不排合计行)
      sheets       - 该簿全部子表名
      error        - 失败原因(ok=False 时)
    """
    res = {"ok": False, "sheet": "", "header_row": 0, "roles": {},
           "source": None, "n_rows": 0, "sheets": [], "error": ""}
    try:
        res["sheets"] = list_sheets(path)
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        try:
            ws = wb[sheet] if (sheet and sheet in wb.sheetnames) else wb[wb.sheetnames[0]]
            hr, col, src = detect_layout_or_shape(ws, log=log)
            res["sheet"] = ws.title
            if not hr:
                res["error"] = "未能识别表头(需含“物料号/编码”列)"
                return res
            res["ok"] = True
            res["header_row"] = hr
            res["roles"] = dict(col)
            res["source"] = src
            res["n_rows"] = max(0, (ws.max_row or hr) - hr)
        finally:
            wb.close()
    except Exception as e:
        res["error"] = str(e)
    return res


def _has_supplier(layout):
    """该表是否带供应商信息（可作供应商来源）。"""
    return "sup_code" in layout["col"] or "sup_name" in layout["col"]


def _has_qty(layout):
    """该表是否带数量/需求数列(物料清单主表的特征)。"""
    return "qty" in layout["col"]


def classify(lay_a, lay_b, n_a=0, n_b=0, log=None):
    """辨识两份表哪份是物料清单(主表)、哪份是供应商明细。返回 ('a'/'b' 为主表, 供应商来源同理)。

    规则(按可靠度依次):
      1) 仅一份带供应商列 -> 它作供应商来源, 另一份作主表(最可靠);
      2) 两份都带供应商列 -> 用"含数量列"区分: 有数量的那份作主表;
      3) 数量列也无法区分 -> 按行数多者作主表, 并 log 明确警告(而非静默择 A);
      4) 两份都不带供应商列 -> 报错提示"未找到供应商列"。
    """
    a_sup, b_sup = _has_supplier(lay_a), _has_supplier(lay_b)
    if not a_sup and not b_sup:
        # 两份都无供应商列: 无法带出供应商, 直接报错让用户确认选错文件
        raise ValueError("两份表都未找到供应商列(供应商代码/名称)，无法确定供应商来源，"
                         "请确认是否选错文件。")
    if a_sup and not b_sup:
        return "b", "a"          # B 无供应商 -> 主表；A 供应商来源
    if b_sup and not a_sup:
        return "a", "b"          # A 无供应商 -> 主表；B 供应商来源
    # 两份都带供应商列: 用数量列存在性区分(物料清单必有数量)
    a_qty, b_qty = _has_qty(lay_a), _has_qty(lay_b)
    if a_qty and not b_qty:
        return "a", "b"
    if b_qty and not a_qty:
        return "b", "a"
    # 数量列也无法区分: 按行数多者作主表, 并明确警告
    if log:
        log("⚠ 两份表都含供应商列且都含/都不含数量列，无法可靠区分主表与供应商表，"
            "已按行数多者(%d vs %d)作主表，请核对结果。" % (n_a, n_b))
    return ("a", "b") if n_a >= n_b else ("b", "a")


# ---------------------------------------------------------------------------
# 输出（复刻客户样本格式）
# ---------------------------------------------------------------------------
_HEAD_FILL = PatternFill("solid", fgColor="FFBDD7EE")     # 新模板表头浅蓝
_HEAD_FONT = Font(name="微软雅黑", size=11, bold=True, color="FF000000")
_DATA_FONT = Font(name="微软雅黑", size=11)                # 新模板数据字体
_TITLE_FONT = Font(name="微软雅黑", size=14, bold=True, color="FF000000")
_THIN = Side(style="thin", color="FF000000")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
# 列宽复刻新模板：A序号 B编码 C名称 D供代 E供信息 其余统一 13
_WIDTHS = [13, 15.83, 39.91, 13, 45.75, 13, 13, 13, 13, 13, 13, 13, 13, 13, 13, 13]

# 输出各列在 OUT_HEADERS 中的位置（1 基），供写值与公式引用，避免散落魔法数字
_C_QTY = 7        # 需求数
_C_RECV = 9       # 实际收货数
_C_LEFT = 12      # 剩余未收数（= 实际收货数 - 需求数）
_C_CASE = 13      # CASE
_C_CASE_QTY = 14  # CASE托数
_C_TEAM = 15      # 班组

_TITLE_ROW_H = 39   # 标题行高
_HEAD_ROW_H = 33    # 表头行高
_DATA_ROW_H = 22    # 数据行统一行高（与参考送货计划一致）


def _sup_sort_key(sc):
    """供应商代码升序排序键：纯数字按数值比，其余按文本比，空/无供应商排最后。

    返回三元组 (分组, 数值, 文本)：0=数字码在前、1=文本码居中、2=空码在末。
    Python 稳定排序，同一供应商内保持原物料清单顺序。
    """
    s = cell_text(sc)
    if not s:
        return (2, 0, "")
    if s.isdigit():
        return (0, int(s), s)
    return (1, 0, s)


def build_plan_sheet(ws, master_rows, sup_map, order_type=None,
                     case_map=None, log=None):
    """把主表行写成送货计划（新模板样式）。

    sup_map    : 归一编码 -> (供应商代码, 供应商名称)。
    order_type : "SUB" / "KD"，统一填入 KD/SUB 列；None 则留空。
    case_map   : 归一编码 -> (CASE, CASE托数, 班组)，来自参考送货计划；None 则不填。

    版式：第 1 行合并 A1:P1 作标题（留空，仅套样式）；第 2 行表头；数据自第 3 行起。
    数据行按供应商代码升序排序后再从 1 编号；每行统一行高（与参考送货计划一致）。
    "剩余未收数"写成 =I{r}-G{r} 公式（实际收货数-需求数），全表加边框、居中。
    返回 (写入行数, 未匹配供应商的编码列表, CASE 命中数)。
    """
    ncol = len(OUT_HEADERS)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    t = ws.cell(1, 1, None)                  # 标题行留空，仅套样式
    t.font = _TITLE_FONT
    t.alignment = _CENTER
    ws.row_dimensions[1].height = _TITLE_ROW_H
    for c in range(1, ncol + 1):            # 表头(第2行)
        cell = ws.cell(2, c, OUT_HEADERS[c - 1])
        cell.font = _HEAD_FONT
        cell.fill = _HEAD_FILL
        cell.alignment = _CENTER
        cell.border = _BORDER
    ws.row_dimensions[2].height = _HEAD_ROW_H

    case_map = case_map or {}
    # 先据供应商查好每行，再按供应商代码升序排序，最后统一编号——
    # 稳定排序保证同一供应商内仍按物料清单原顺序。
    recs = []
    missing = []
    for row in master_rows:
        code = norm_code(row["code"])
        sup = sup_map.get(code)
        if sup is None:
            missing.append(code)
        sc, sn = (sup if sup else (None, None))
        recs.append({"row": row, "sc": sc, "sn": sn, "cinfo": case_map.get(code)})
    recs.sort(key=lambda rc: _sup_sort_key(rc["sc"]))

    hit_case = 0
    r = 3
    for i, rc in enumerate(recs, 1):
        row, sc, sn = rc["row"], rc["sc"], rc["sn"]
        cinfo = rc["cinfo"]
        if cinfo:
            hit_case += 1
        ca, cq, tm = (cinfo if cinfo else (None, None, None))
        vals = [i, row["code"], row.get("cname"), sc, sn, order_type or None,
                row.get("qty"), None, None, None, None,
                "=%s%d-%s%d" % (get_column_letter(_C_RECV), r,
                                get_column_letter(_C_QTY), r),
                ca, cq, tm, None]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v)
            cell.font = _DATA_FONT
            cell.alignment = _CENTER
            cell.border = _BORDER
        ws.row_dimensions[r].height = _DATA_ROW_H
        r += 1

    for c, w in enumerate(_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    if log and missing:
        log("有 %d 个物料在供应商明细中未找到供应商，已留空：%s%s"
            % (len(missing), "、".join(missing[:8]), " 等" if len(missing) > 8 else ""))
    return r - 3, missing, hit_case


def build_supplier_map(sup_rows, log=None):
    """从供应商明细行建 归一编码 -> (供应商代码, 供应商名称) 映射。

    一码多供应商时以首见为准，并在日志里提示冲突数（避免静默择一）。
    """
    m = {}
    conflicts = 0
    for row in sup_rows:
        code = norm_code(row["code"])
        if not code:
            continue
        pair = (cell_text(row.get("sup_code")) or None,
                cell_text(row.get("sup_name")) or None)
        if code in m:
            if m[code] != pair:
                conflicts += 1
            continue
        m[code] = pair
    if log and conflicts:
        log("注意：供应商明细中有 %d 个物料存在多个不同供应商，已取首个。" % conflicts)
    return m


# ---------------------------------------------------------------------------
# 参考送货计划：按物料编码带出 CASE / CASE托数 / 班组
# ---------------------------------------------------------------------------
_REF_KEYS = {"code": ["物料编码", "物料号", "零部件代码", "编码"],
             "case": ["CASE"], "case_qty": ["CASE托数", "托数"],
             "team": ["班组"]}


def _match_ref_header(ws, scan_rows=8):
    """在参考表前若干行找含"物料编码 + CASE"的明细表头行，返回 (行号, {角色:列号})。

    参考表常有多个 sheet（如透视 Sheet2 + 明细"零件到货计划"）。透视汇总表虽也含
    "班组"、且"计数项:物料编码"会被物料编码子串命中，但它没有 CASE 列——故以 CASE 为
    判据可稳妥排除透视表。code 匹配优先精确表头，避免误取"计数项:物料编码"之类。
    找不到返回 (None, {})。
    """
    for hr in range(1, min(scan_rows, ws.max_row) + 1):
        col = {}
        for c in range(1, ws.max_column + 1):
            txt = cell_text(ws.cell(hr, c).value)
            if not txt:
                continue
            for role, keys in _REF_KEYS.items():
                if role in col:
                    continue
                # code 只认精确表头（防"计数项:物料编码"透视列）；其余可含匹配
                if role == "code":
                    if any(k == txt for k in keys):
                        col[role] = c
                elif any(k == txt for k in keys) or any(k in txt for k in keys):
                    col[role] = c
        # CASE 是明细表独有列，作硬判据；再要求有 班组 或 托数 之一
        if "code" in col and "case" in col and ("team" in col or "case_qty" in col):
            return hr, col
    return None, {}


def build_case_map(path, log=None):
    """从参考送货计划读 归一编码 -> (CASE, CASE托数, 班组)。一码多值以首见为准。

    自动跳过透视/汇总 sheet，找含物料编码且带 CASE/班组 的明细 sheet。
    读不到（文件缺失/无合适表头）返回空 dict，不报错——参考表本就是可选项。
    """
    def _lg(msg):
        if log:
            log(msg)
    if not path:
        return {}
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        _lg("参考送货计划读取失败，已跳过 CASE/班组：%s" % e)
        return {}
    try:
        for name in wb.sheetnames:
            ws = wb[name]
            hr, col = _match_ref_header(ws)
            if not hr:
                continue
            m = {}
            for r in range(hr + 1, (ws.max_row or hr) + 1):
                code = norm_code(ws.cell(r, col["code"]).value)
                if not code or code in m:
                    continue
                ca = cell_text(ws.cell(r, col["case"]).value) if "case" in col else ""
                cq = ws.cell(r, col["case_qty"]).value if "case_qty" in col else None
                tm = cell_text(ws.cell(r, col["team"]).value) if "team" in col else ""
                if ca or cq is not None or tm:
                    m[code] = (ca or None, cq, tm or None)
            _lg("参考送货计划：从工作表「%s」读到 %d 条 CASE/班组 记录。" % (name, len(m)))
            return m
        _lg("参考送货计划里未找到含「物料编码 + CASE/班组」的表，已跳过。")
        return {}
    finally:
        wb.close()


def run(file_a, file_b, sheet_a=None, sheet_b=None, out_dir=None, log=None,
        order_type=None, ref_plan=None):
    """送货计划表制作主流程。两份输入顺序任意，自动辨识主表/供应商来源。

    返回 dict：{plan_path, out_dir, rows, matched, missing, master_file, supplier_file}。
    """
    def _lg(msg):
        if log:
            log(msg)

    # 物料清单"需求数/数量"若是未刷新的公式,data_only 读为 None → 送货计划需求数
    # 静默留空。入口两份表都先查一遍(此刻还没辨识主表),把隐患变成可见提示。
    from . import common_core
    common_core.warn_if_uncached(file_a, _lg, sheet_a, what="需求数/数量")
    common_core.warn_if_uncached(file_b, _lg, sheet_b, what="需求数/数量")

    rows_a, lay_a = load_sheet(file_a, sheet_a, log=_lg)
    rows_b, lay_b = load_sheet(file_b, sheet_b, log=_lg)
    master_key, sup_key = classify(lay_a, lay_b, len(rows_a), len(rows_b), log=_lg)
    pack = {"a": (rows_a, lay_a, file_a), "b": (rows_b, lay_b, file_b)}
    master_rows, _lm, master_file = pack[master_key]
    sup_rows, _ls, sup_file = pack[sup_key]
    _lg("主表(物料清单)：%s —— %d 行" % (os.path.basename(master_file), len(master_rows)))
    _lg("供应商来源：%s —— %d 行" % (os.path.basename(sup_file), len(sup_rows)))

    sup_map = build_supplier_map(sup_rows, log=_lg)

    ot = (order_type or "").strip().upper() or None
    if ot:
        _lg("KD/SUB 列统一填：%s" % ot)
    case_map = build_case_map(ref_plan, log=_lg) if ref_plan else {}

    if out_dir is None:
        st = settings_mod.get_settings()
        out_dir = _paths.resolve_output_dir("delivery", **st.output_kwargs())
    else:
        os.makedirs(out_dir, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    n, missing, hit_case = build_plan_sheet(
        ws, master_rows, sup_map, order_type=ot, case_map=case_map, log=_lg)
    matched = n - len(missing)
    _lg("已生成 %d 行，供应商匹配 %d / %d。" % (n, matched, n))
    if case_map:
        _lg("CASE/班组 按物料编码匹配 %d / %d 行。" % (hit_case, n))

    plan_path = os.path.join(out_dir, "送货计划.xlsx")
    try:
        wb.save(plan_path)
    except PermissionError:
        raise PermissionError("无法保存 %s —— 请先在 Excel 里关闭该文件后重试" % plan_path)
    _lg("已生成送货计划：%s" % plan_path)
    return {"plan_path": plan_path, "out_dir": out_dir, "rows": n,
            "matched": matched, "missing": missing, "order_type": ot,
            "case_hit": hit_case, "case_used": bool(case_map),
            "master_file": master_file, "supplier_file": sup_file}
